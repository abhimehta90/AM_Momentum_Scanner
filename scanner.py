"""
AM_Watch Technical Scanner v2
=============================
Scans the full Nifty 500 universe against a rules-based framework
calibrated to the user's *custom* TradingView indicator settings:

    BB %b :   length=50, source=ohlc4, stdDev=2
    MACD  :   fast=3, slow=21, signal=9, source=close
    ADX   :   DI length=14 (Wilder), ADX smoothing=SMA(14) on DX

Produces a color-coded Excel workbook AND a self-contained HTML
dashboard with hover tooltips (raw indicator values), an "add from
Nifty 500" dropdown, and per-row remove buttons. The user's watchlist
is persisted in the browser via localStorage so it survives file
re-opens.

Run:      python3 scanner.py
"""

import os, sys, json, warnings, datetime as dt
from pathlib import Path
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
import yfinance as yf

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
BASE_DIR      = Path(__file__).resolve().parent
NIFTY_CSV     = BASE_DIR / "nifty500.csv"
# Output dir can be overridden via env var (e.g. SCANNER_OUT_DIR=./site in CI).
# Falls back to the Cowork workspace for local runs.
_DEFAULT_OUT  = "/sessions/wonderful-nifty-darwin/mnt/AKM"
OUT_DIR       = Path(os.environ.get("SCANNER_OUT_DIR", _DEFAULT_OUT))
if not OUT_DIR.parent.exists():
    OUT_DIR = BASE_DIR / "site"
SNAPSHOT_DIR  = BASE_DIR / ".scanner_snapshots"
OUT_DIR.mkdir(parents=True, exist_ok=True)
SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
STAMP         = dt.datetime.now().strftime("%Y-%m-%d")
STAMP_HUMAN   = dt.datetime.now().strftime("%A, %-d %b %Y")

# ATR stop loss multiplier (close - ATR_MULT * ATR)
ATR_MULT = 2.0
# Strong Buy scan threshold for the "Nifty 500 scan" tab (>= this score)
SB_SCAN_MIN_SCORE = 70

# Default watchlist seeded into the HTML (user can add/remove in-browser)
DEFAULT_WATCHLIST = [
    "CUPID", "CRAFTSMAN", "SHRIPISTON", "ASHOKLEY", "ATHERENERG",
    "PRICOLLTD", "SAGILITY", "TITAN", "TVSMOTOR", "CUB",
    "TORNTPHARM", "MARUTI", "LAURUSLABS", "SHRIRAMFIN", "LTF",
    "TATACAP", "SOUTHBANK", "FEDERALBNK", "UJJIVANSFB", "SBIN",
    "AUBANK", "MAHABANK",
]

# Point-weighted signals.
# Raw group max (BB=26, MACD=24, ADX=50) then normalized and scaled by
# GROUP_WEIGHT (BB 30, MACD 40, ADX 30) → composite on 0–100 scale.
SIGNAL_WEIGHTS = {
    # ----- BB group (max 26) -----
    "BB_LT_uptrend":      6,   # Close > 50-day MA
    "BB_ST_uptrend":      8,   # BB% rising 3 days
    "BB_val_strong":     12,   # BB% ≥ 0.8 — HIGHEST (mutex w/ bottom zone)
    "BB_bottom_zone":     8,   # BB% 0.1–0.3 (mutex w/ val_strong)
    # overbought penalty removed per user spec (weight = 0, not displayed)

    # ----- MACD group (max 24) — CROSSOVER CENTRIC -----
    # All "cross" buckets below are mutually exclusive: a stock has
    # exactly one crossover state at any time.
    "MACD_3d_up":                      4,   # line rising 3 days
    "MACD_cross_above_zero_recent":   20,   # crossed 1–5d ago, MACD ≥ 0 → STRONGEST
    "MACD_cross_near_zero_recent":    14,   # crossed 1–5d ago, MACD < 0 but close
    "MACD_cross_far_below_recent":     4,   # crossed 1–5d ago, far below 0 → muted
    "MACD_cross_sustained":            2,   # cross 6–15d ago (muted / decaying)
    "MACD_closing_in":                 4,   # hist rising, still <0 (pre-cross)

    # ----- ADX group (max 50) -----
    "ADX_reversal":      12,   # rising 2–5d after declining — STRONG
    "ADX_continuing":     6,   # rising ≥6d — normal (mutex w/ reversal)
    "ADX_above_18":       5,
    "DI+_rising":         8,   # user explicit
    "DI+_uptrend":        8,   # user explicit
    "DI_gap_wide":        5,   # DI+ − DI- ≥ 10
    "DI_gap_widening":   12,   # gap widening 2–5d — STRONG
}

SIG_BY_GROUP = {
    "BB":   ["BB_LT_uptrend", "BB_ST_uptrend", "BB_val_strong", "BB_bottom_zone"],
    "MACD": ["MACD_3d_up", "MACD_cross_above_zero_recent",
             "MACD_cross_near_zero_recent", "MACD_cross_far_below_recent",
             "MACD_cross_sustained", "MACD_closing_in"],
    "ADX":  ["ADX_reversal", "ADX_continuing", "ADX_above_18",
             "DI+_rising", "DI+_uptrend", "DI_gap_wide", "DI_gap_widening"],
}

# Raw theoretical maxes (sum of per-group best achievable, accounting for mutex)
GROUP_MAX = {
    "BB":   6 + 8 + 12,               # LT + 3d + val_strong = 26
    "MACD": 4 + 20,                   # 3d↑ + cross_above_zero_recent = 24
    "ADX":  12 + 5 + 8 + 8 + 5 + 12,  # rev+≥18+DI+r+DI+u+gap+gapWid = 50
}

# Overall group weights (percent of composite) per user spec
GROUP_WEIGHT = {"BB": 30, "MACD": 40, "ADX": 30}
TOTAL_MAX = sum(GROUP_WEIGHT.values())  # 100

# Kept for backward-compat in HTML sub-label
WEIGHTS = GROUP_WEIGHT

# Score → category thresholds (high to low)
CATEGORIES = [
    (75, "Strong Buy"),
    (50, "Watch"),
    (25, "Neutral"),
    (0,  "Avoid"),
]

# Indicator params — matched to user's TradingView setup
BB_LEN = 50
BB_DEV = 2
MACD_FAST, MACD_SLOW, MACD_SIG = 3, 21, 9
ADX_DI_LEN = 14          # Wilder RMA for +DM/-DM/TR
ADX_SMOOTH_LEN = 14      # SMA length on DX (matches user's ADX v4)


# ---------------------------------------------------------------------------
# NIFTY 500 LOAD
# ---------------------------------------------------------------------------
def load_nifty500() -> pd.DataFrame:
    df = pd.read_csv(NIFTY_CSV)
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={"Company Name": "name", "Symbol": "ticker",
                            "Industry": "industry"})
    return df[["ticker", "name", "industry"]].copy()


# ---------------------------------------------------------------------------
# BATCH DATA FETCH
# ---------------------------------------------------------------------------
def fetch_batch(tickers: list[str]) -> dict[str, pd.DataFrame]:
    """
    Batch-download daily OHLCV for many tickers at once. For any ticker
    whose most recent bar has OHL but a NaN close (Yahoo Finance common
    delay for NSE), patch the close from Ticker.fast_info.last_price so
    the latest session is included in indicator calculations.
    """
    symbols = [f"{t}.NS" for t in tickers]
    print(f"Fetching {len(symbols)} tickers from Yahoo Finance…")
    # 3y period gives ~750 daily bars (plenty) AND ~150 weekly bars after
    # resampling (enough for BB50 / MACD 3-21-9 / ADX14 on weekly timeframe).
    raw = yf.download(
        symbols, period="3y", interval="1d", group_by="ticker",
        progress=False, auto_adjust=False, threads=True,
    )

    # First pass: collect raw frames
    frames: dict[str, pd.DataFrame] = {}
    needs_patch: list[str] = []
    for t, sym in zip(tickers, symbols):
        try:
            d = raw[sym].copy()
            if d.empty:
                continue
            # If the last row has OHL but NaN close, flag for patching
            last = d.iloc[-1]
            has_ohl = all(pd.notna(last.get(k)) for k in ["Open","High","Low"])
            if has_ohl and pd.isna(last.get("Close")):
                needs_patch.append(t)
            frames[t] = d
        except (KeyError, AttributeError):
            pass

    # Second pass: patch NaN closes with fast_info.last_price
    if needs_patch:
        print(f"  Patching {len(needs_patch)} tickers with latest live close…")
        patched = 0
        for t in needs_patch:
            try:
                fi = yf.Ticker(f"{t}.NS").fast_info
                lp = fi.get("last_price") or fi.get("lastPrice")
                if lp is None or not (lp > 0):
                    continue
                d = frames[t]
                last_idx = d.index[-1]
                d.loc[last_idx, "Close"] = float(lp)
                if "Adj Close" in d.columns:
                    d.loc[last_idx, "Adj Close"] = float(lp)
                frames[t] = d
                patched += 1
            except Exception:
                pass
        print(f"  Patched {patched}/{len(needs_patch)}")

    # Final: dropna on Close and filter by history length
    out: dict[str, pd.DataFrame] = {}
    for t, d in frames.items():
        d = d.dropna(subset=["Close"]).copy()
        if len(d) >= 70:
            out[t] = d
    print(f"  → usable history: {len(out)}/{len(tickers)}")
    return out


# ---------------------------------------------------------------------------
# INDICATORS — matched to custom TradingView settings
# ---------------------------------------------------------------------------
def rma(series: pd.Series, length: int) -> pd.Series:
    """Wilder's moving average (RMA)."""
    return series.ewm(alpha=1/length, adjust=False).mean()

def compute_indicators(df: pd.DataFrame) -> pd.DataFrame:
    o, h, l, c = [df[k].astype(float) for k in ["Open", "High", "Low", "Close"]]

    # ---- BB %b (length=50, source=ohlc4, stdDev=2) ----
    ohlc4 = (o + h + l + c) / 4
    basis = ohlc4.rolling(BB_LEN).mean()
    dev   = ohlc4.rolling(BB_LEN).std(ddof=0) * BB_DEV
    upper, lower = basis + dev, basis - dev
    df["BB_pct"] = (ohlc4 - lower) / (upper - lower)
    df["BB_basis"] = basis
    df["BB_upper"] = upper
    df["BB_lower"] = lower

    # ---- MACD (3, 21, 9) on close ----
    ema_f = c.ewm(span=MACD_FAST, adjust=False).mean()
    ema_s = c.ewm(span=MACD_SLOW, adjust=False).mean()
    df["MACD"]      = ema_f - ema_s
    df["MACD_sig"]  = df["MACD"].ewm(span=MACD_SIG, adjust=False).mean()
    df["MACD_hist"] = df["MACD"] - df["MACD_sig"]

    # ---- ADX (DI Wilder 14, ADX = EMA(20) on DX) ----
    up, dn = h.diff(), -l.diff()
    plus_dm  = ((up > dn) & (up > 0))  * up
    minus_dm = ((dn > up) & (dn > 0)) * dn
    tr = pd.concat([h - l,
                    (h - c.shift()).abs(),
                    (l - c.shift()).abs()], axis=1).max(axis=1)
    atr       = rma(tr, ADX_DI_LEN)
    df["DI+"] = 100 * rma(plus_dm,  ADX_DI_LEN) / atr
    df["DI-"] = 100 * rma(minus_dm, ADX_DI_LEN) / atr
    dx        = 100 * (df["DI+"] - df["DI-"]).abs() / (df["DI+"] + df["DI-"])
    df["ADX"] = dx.rolling(ADX_SMOOTH_LEN).mean()

    # Store ATR for stop loss calc (reuse Wilder ATR already computed above)
    df["ATR"] = atr

    # 50-day MA for LT trend context
    df["MA50"] = c.rolling(50).mean()
    return df


def resample_weekly(df: pd.DataFrame) -> pd.DataFrame:
    """Resample daily OHLCV to weekly (W-FRI) bars."""
    agg = {"Open": "first", "High": "max", "Low": "min",
           "Close": "last", "Volume": "sum"}
    w = df.resample("W-FRI").agg(agg).dropna(subset=["Close"])
    if "Adj Close" in df.columns:
        w["Adj Close"] = df["Adj Close"].resample("W-FRI").last()
    return w


def _strict_uptrend(series: pd.Series, days: int) -> bool:
    t = series.dropna().tail(days + 1).values
    if len(t) < days + 1:
        return False
    return all(t[i] > t[i - 1] for i in range(1, len(t)))


# ---------------------------------------------------------------------------
# SCORING
# ---------------------------------------------------------------------------
def score_ticker(df: pd.DataFrame) -> dict:
    last = df.iloc[-1]
    close, bb_pct = df["Close"], df["BB_pct"]
    macd, sig_ln, hist = df["MACD"], df["MACD_sig"], df["MACD_hist"]
    adx, dip, dim = df["ADX"], df["DI+"], df["DI-"]

    sigs: dict[str, bool] = {s: False for grp in SIG_BY_GROUP.values() for s in grp}

    # ============================================================
    # BB group
    # ============================================================
    sigs["BB_LT_uptrend"]  = bool(last["Close"] > last["MA50"])
    sigs["BB_ST_uptrend"]  = _strict_uptrend(bb_pct, 3)
    sigs["BB_val_strong"]  = bool(last["BB_pct"] >= 0.8)
    # bottom_zone mutex w/ val_strong
    sigs["BB_bottom_zone"] = bool(
        0.1 <= last["BB_pct"] <= 0.3 and not sigs["BB_val_strong"]
    )

    # ============================================================
    # MACD group — crossover-centric
    # ============================================================
    sigs["MACD_3d_up"] = _strict_uptrend(macd, 3)

    # Find the most recent bullish crossover: hist transitions ≤0 → >0.
    # We scan hist (MACD-Signal) for the last sign flip to positive.
    hist_vals = hist.values
    days_since_cross: int | None = None
    cross_idx: int | None = None
    for i in range(len(hist_vals) - 1, 0, -1):
        if not np.isnan(hist_vals[i]) and not np.isnan(hist_vals[i - 1]):
            if hist_vals[i] > 0 and hist_vals[i - 1] <= 0:
                cross_idx = i
                days_since_cross = (len(hist_vals) - 1) - i
                break

    # 60-day |MACD| range for "far below vs close to zero" judgment
    rng_series = macd.abs().rolling(60).max()
    rng = float(rng_series.iloc[-1]) if not np.isnan(rng_series.iloc[-1]) else 0.0

    if days_since_cross is not None and cross_idx is not None:
        macd_at_cross = float(macd.iloc[cross_idx])
        if days_since_cross <= 5:
            # Recent crossover — categorize by position vs zero line
            if macd_at_cross >= 0:
                sigs["MACD_cross_above_zero_recent"] = True
            elif rng > 0 and abs(macd_at_cross) / rng < 0.25:
                sigs["MACD_cross_near_zero_recent"] = True
            else:
                sigs["MACD_cross_far_below_recent"] = True
        elif days_since_cross <= 15:
            # 6–15 days ago — sustained (muted)
            sigs["MACD_cross_sustained"] = True
        # >15 days ago → no crossover signal (neutral per user spec)

    any_cross_active = any(
        sigs[k] for k in (
            "MACD_cross_above_zero_recent",
            "MACD_cross_near_zero_recent",
            "MACD_cross_far_below_recent",
            "MACD_cross_sustained",
        )
    )

    # closing_in: histogram rising but still <0 AND no active crossover
    sigs["MACD_closing_in"] = bool(
        _strict_uptrend(hist, 2)
        and last["MACD_hist"] < 0
        and not any_cross_active
    )

    # ============================================================
    # ADX group
    # ============================================================
    # Count rising-streak length at the tail of ADX
    adx_clean = adx.dropna().values
    rising_streak = 0
    for i in range(len(adx_clean) - 1, 0, -1):
        if adx_clean[i] > adx_clean[i - 1]:
            rising_streak += 1
        else:
            break

    # Reversal: 2–5 rising days preceded by a declining period
    if 2 <= rising_streak <= 5 and len(adx_clean) >= rising_streak + 4:
        prior = adx_clean[-rising_streak - 4 : -rising_streak]
        sigs["ADX_reversal"] = bool(prior[-1] < prior[0])
    # Continuing: ≥6 rising days (mutex w/ reversal)
    if rising_streak >= 6:
        sigs["ADX_continuing"] = True

    sigs["ADX_above_18"] = bool(last["ADX"] >= 18)

    dip_prev = dip.iloc[-3] if len(dip) >= 3 else dip.iloc[0]
    sigs["DI+_rising"] = bool(
        (last["DI+"] >= 20 and last["DI+"] > dip_prev) or
        (last["DI+"] < 20 and _strict_uptrend(dip, 2))
    )
    sigs["DI+_uptrend"] = _strict_uptrend(dip, 3)
    sigs["DI_gap_wide"] = bool((last["DI+"] - last["DI-"]) >= 10)

    # DI gap widening: gap rising ≥2 consecutive days within last 5
    gap_series = (dip - dim).dropna()
    gap_widening = False
    if len(gap_series) >= 3:
        diffs = np.diff(gap_series.tail(6).values)
        streak = 0
        for d in reversed(diffs):
            if d > 0:
                streak += 1
            else:
                break
        gap_widening = streak >= 2
    sigs["DI_gap_widening"] = bool(gap_widening)

    # ============================================================
    # Composite score
    # ============================================================
    bb_raw   = sum(SIGNAL_WEIGHTS[s] for s in SIG_BY_GROUP["BB"]   if sigs[s])
    macd_raw = sum(SIGNAL_WEIGHTS[s] for s in SIG_BY_GROUP["MACD"] if sigs[s])
    adx_raw  = sum(SIGNAL_WEIGHTS[s] for s in SIG_BY_GROUP["ADX"]  if sigs[s])

    bb_pct_s   = min(1.0, max(0.0, bb_raw   / GROUP_MAX["BB"]))
    macd_pct_s = min(1.0, max(0.0, macd_raw / GROUP_MAX["MACD"]))
    adx_pct_s  = min(1.0, max(0.0, adx_raw  / GROUP_MAX["ADX"]))

    composite = (bb_pct_s   * GROUP_WEIGHT["BB"]   +
                 macd_pct_s * GROUP_WEIGHT["MACD"] +
                 adx_pct_s  * GROUP_WEIGHT["ADX"])
    cat = next(label for th, label in CATEGORIES if composite >= th)

    # ATR-based stop loss
    atr_last  = float(last["ATR"]) if pd.notna(last.get("ATR")) else 0.0
    close_px  = float(last["Close"])
    stop_px   = close_px - ATR_MULT * atr_last if atr_last > 0 else 0.0
    stop_pct  = (stop_px - close_px) / close_px * 100 if stop_px > 0 else 0.0

    return {
        "close":     round(close_px, 2),
        "bb_pct":    round(float(last["BB_pct"]), 3),
        "bb_basis":  round(float(last["BB_basis"]), 2),
        "bb_upper":  round(float(last["BB_upper"]), 2),
        "bb_lower":  round(float(last["BB_lower"]), 2),
        "macd":      round(float(last["MACD"]), 2),
        "macd_sig":  round(float(last["MACD_sig"]), 2),
        "macd_hist": round(float(last["MACD_hist"]), 2),
        "adx":       round(float(last["ADX"]), 2),
        "di_plus":   round(float(last["DI+"]), 2),
        "di_minus":  round(float(last["DI-"]), 2),
        "ma50":      round(float(last["MA50"]), 2),
        "atr":       round(atr_last, 2),
        "stop":      round(stop_px, 2),
        "stop_pct":  round(stop_pct, 2),
        "bb_raw":     bb_raw,
        "macd_raw":   macd_raw,
        "adx_raw":    adx_raw,
        "bb_score":   round(bb_pct_s * 100),
        "macd_score": round(macd_pct_s * 100),
        "adx_score":  round(adx_pct_s * 100),
        "score":      round(composite, 1),
        "category":   cat,
        "days_since_cross": -1 if days_since_cross is None else int(days_since_cross),
        **{k: int(v) for k, v in sigs.items()},
    }


# ---------------------------------------------------------------------------
# SNAPSHOT (day-over-day delta)
# ---------------------------------------------------------------------------
def load_prev_snapshot() -> dict:
    """Load most recent snapshot prior to today (for day-over-day delta)."""
    try:
        files = sorted(p for p in SNAPSHOT_DIR.glob("*.json")
                       if p.stem < STAMP)
        if not files:
            return {}
        prev = files[-1]
        print(f"Loaded prior snapshot: {prev.name}")
        return json.loads(prev.read_text())
    except Exception as e:
        print(f"  (no prior snapshot: {e})")
        return {}


def save_snapshot(df: pd.DataFrame) -> None:
    """Save today's scores keyed by ticker for tomorrow's delta calc."""
    snap = {}
    ok = df[df["status"] == "ok"]
    for _, r in ok.iterrows():
        snap[r["ticker"]] = {
            "score":    float(r["score"]),
            "category": r["category"],
            "close":    float(r["close"]),
        }
    out = SNAPSHOT_DIR / f"{STAMP}.json"
    out.write_text(json.dumps(snap, separators=(",", ":")))
    print(f"Saved snapshot: {out.name}")


# ---------------------------------------------------------------------------
# MARKET BREADTH
# ---------------------------------------------------------------------------
def compute_breadth(df: pd.DataFrame, data: dict) -> dict:
    """Nifty 500 advance/decline + % in Strong Buy + % above 50d MA + NIFTY 50."""
    ok = df[df["status"] == "ok"]
    n_total = len(ok)
    n_adv = int((ok["close"] > ok["ma50"]).sum()) if "ma50" in ok.columns else 0
    n_dec = n_total - n_adv
    n_sb  = int((ok["category"] == "Strong Buy").sum())
    n_wa  = int((ok["category"] == "Watch").sum())

    # 1-day % change — compare last close to prior close for each ticker
    day_up = day_dn = 0
    day_changes = []
    for t in ok["ticker"]:
        if t in data and len(data[t]) >= 2:
            c0, c1 = data[t]["Close"].iloc[-2], data[t]["Close"].iloc[-1]
            if pd.notna(c0) and pd.notna(c1) and c0 > 0:
                chg = (c1 - c0) / c0 * 100
                day_changes.append(chg)
                if chg > 0: day_up += 1
                elif chg < 0: day_dn += 1

    avg_chg = float(np.mean(day_changes)) if day_changes else 0.0

    # NIFTY 50 (^NSEI)
    nifty_lvl = nifty_chg = None
    try:
        nidx = yf.download("^NSEI", period="5d", interval="1d",
                           progress=False, auto_adjust=False)
        if not nidx.empty and len(nidx) >= 2:
            last = float(nidx["Close"].iloc[-1])
            prev = float(nidx["Close"].iloc[-2])
            nifty_lvl = round(last, 2)
            nifty_chg = round((last - prev) / prev * 100, 2)
    except Exception as e:
        print(f"  (NIFTY fetch failed: {e})")

    return {
        "n_total":   n_total,
        "n_adv":     n_adv,
        "n_dec":     n_dec,
        "adv_pct":   round(n_adv  / n_total * 100, 1) if n_total else 0.0,
        "day_up":    day_up,
        "day_dn":    day_dn,
        "avg_chg":   round(avg_chg, 2),
        "n_sb":      n_sb,
        "sb_pct":    round(n_sb / n_total * 100, 1) if n_total else 0.0,
        "n_wa":      n_wa,
        "nifty":     nifty_lvl,
        "nifty_chg": nifty_chg,
    }


# ---------------------------------------------------------------------------
# SCAN
# ---------------------------------------------------------------------------
def _weekly_score(df_daily: pd.DataFrame) -> dict:
    """Resample to weekly and run scoring. Returns subset with 'w_' prefix
    or empty dict if not enough history."""
    try:
        w = resample_weekly(df_daily[["Open","High","Low","Close","Volume"] +
                                      (["Adj Close"] if "Adj Close" in df_daily.columns else [])])
        if len(w) < 60:
            return {}
        w = compute_indicators(w)
        s = score_ticker(w)
        return {
            "w_score":    s["score"],
            "w_category": s["category"],
            "w_bb_pct":   s["bb_pct"],
            "w_adx":      s["adx"],
            "w_di_plus":  s["di_plus"],
            "w_macd":     s["macd"],
            "w_macd_sig": s["macd_sig"],
            "w_macd_hist":s["macd_hist"],
            "w_bb_raw":    s["bb_raw"],
            "w_macd_raw":  s["macd_raw"],
            "w_adx_raw":   s["adx_raw"],
            "w_bb_score":  s["bb_score"],
            "w_macd_score":s["macd_score"],
            "w_adx_score": s["adx_score"],
            "w_dsc":      s["days_since_cross"],
            # keep weekly signal booleans under distinct keys
            **{f"w_{k}": s[k] for k in
               [s0 for grp in SIG_BY_GROUP.values() for s0 in grp]},
        }
    except Exception:
        return {}


def scan_universe(universe: pd.DataFrame) -> tuple[pd.DataFrame, dict, dict]:
    tickers = universe["ticker"].tolist()
    data = fetch_batch(tickers)
    rows = []
    print("Computing indicators (daily + weekly)…")
    for _, u in universe.iterrows():
        t = u["ticker"]
        if t not in data:
            rows.append({"ticker": t, "name": u["name"],
                         "industry": u["industry"], "status": "no data"})
            continue
        try:
            d = compute_indicators(data[t].copy())
            daily = score_ticker(d)
            weekly = _weekly_score(data[t])
            rows.append({"ticker": t, "name": u["name"],
                         "industry": u["industry"], "status": "ok",
                         **daily, **weekly})
        except Exception as e:
            rows.append({"ticker": t, "name": u["name"],
                         "industry": u["industry"], "status": f"err: {e}"})
    df = pd.DataFrame(rows)

    # Day-over-day delta
    prev = load_prev_snapshot()
    if prev:
        def delta(row):
            t = row["ticker"]
            p = prev.get(t)
            if not p or row.get("status") != "ok":
                return None
            return round(float(row["score"]) - float(p["score"]), 1)
        def prev_cat(row):
            t = row["ticker"]
            p = prev.get(t)
            return (p or {}).get("category")
        df["delta"] = df.apply(delta, axis=1)
        df["prev_category"] = df.apply(prev_cat, axis=1)
    else:
        df["delta"] = None
        df["prev_category"] = None

    breadth = compute_breadth(df, data)
    return df, data, breadth


# ---------------------------------------------------------------------------
# EXCEL OUTPUT (watchlist only)
# ---------------------------------------------------------------------------
RULE_COLS = [
    ("BB LT Up",         "BB_LT_uptrend",                  "BB"),
    ("BB 3d Up",         "BB_ST_uptrend",                  "BB"),
    ("BB ≥0.8",          "BB_val_strong",                  "BB"),
    ("BB 0.1–0.3",       "BB_bottom_zone",                 "BB"),
    ("MACD 3d Up",       "MACD_3d_up",                     "MACD"),
    ("Cross Above 0 (1–5d)", "MACD_cross_above_zero_recent",   "MACD"),
    ("Cross Near 0 (1–5d)",  "MACD_cross_near_zero_recent",    "MACD"),
    ("Cross Far Below (1–5d)","MACD_cross_far_below_recent",   "MACD"),
    ("Cross 6–15d (sust.)",  "MACD_cross_sustained",           "MACD"),
    ("MACD Pre-Cross",   "MACD_closing_in",                "MACD"),
    ("ADX Reversal",     "ADX_reversal",                   "ADX"),
    ("ADX Continuing",   "ADX_continuing",                 "ADX"),
    ("ADX ≥18",          "ADX_above_18",                   "ADX"),
    ("DI+ Rising",       "DI+_rising",                     "ADX"),
    ("DI+ Uptrend",      "DI+_uptrend",                    "ADX"),
    ("DI Gap Wide",      "DI_gap_wide",                    "ADX"),
    ("DI Gap Widening",  "DI_gap_widening",                "ADX"),
]

GREEN  = PatternFill("solid", start_color="C6EFCE")
RED    = PatternFill("solid", start_color="FFC7CE")
HEAD   = PatternFill("solid", start_color="305496")
CAT_FILLS = {
    "Strong Buy": PatternFill("solid", start_color="63BE7B"),
    "Watch":      PatternFill("solid", start_color="FFEB84"),
    "Neutral":    PatternFill("solid", start_color="F0F0F0"),
    "Avoid":      PatternFill("solid", start_color="F8696B"),
}

def write_excel(df: pd.DataFrame, path: Path, watchlist: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Scan {STAMP}"

    ws["A1"] = f"AM_Watch Scanner v2 — {STAMP} — Custom BB/MACD/ADX"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = HEAD
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=11 + len(RULE_COLS))
    ws.row_dimensions[1].height = 22

    ws["A2"] = (f"BB %b(50,ohlc4,2)  |  MACD(3,21,9)  |  "
                f"ADX(DI=14, smooth=SMA14)  |  "
                f"Group weights BB={GROUP_WEIGHT['BB']}% "
                f"MACD={GROUP_WEIGHT['MACD']}% ADX={GROUP_WEIGHT['ADX']}%")
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=11 + len(RULE_COLS))

    headers = ["Ticker", "Name", "Close", "Category", "Score",
               "BB%", "MACD", "MACD-Sig", "Hist", "ADX", "DI+/-"
               ] + [h for h, _, _ in RULE_COLS]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[4].height = 38

    df_w = df[df["ticker"].isin(watchlist) & (df["status"] == "ok")
              ].sort_values("score", ascending=False)

    row = 5
    for _, r in df_w.iterrows():
        ws.cell(row=row, column=1, value=r["ticker"]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=r["name"][:32])
        ws.cell(row=row, column=3, value=r["close"]).number_format = "#,##0.00"
        cat = ws.cell(row=row, column=4, value=r["category"])
        cat.fill = CAT_FILLS.get(r["category"], PatternFill())
        cat.font = Font(bold=True)
        cat.alignment = Alignment(horizontal="center")
        sc = ws.cell(row=row, column=5, value=r["score"])
        sc.number_format = "0.0"
        sc.font = Font(bold=True)
        ws.cell(row=row, column=6,  value=r["bb_pct"]).number_format = "0.000"
        ws.cell(row=row, column=7,  value=r["macd"]).number_format = "0.00"
        ws.cell(row=row, column=8,  value=r["macd_sig"]).number_format = "0.00"
        ws.cell(row=row, column=9,  value=r["macd_hist"]).number_format = "0.00"
        ws.cell(row=row, column=10, value=r["adx"]).number_format = "0.00"
        ws.cell(row=row, column=11,
                value=f"{r['di_plus']:.1f} / {r['di_minus']:.1f}"
                ).alignment = Alignment(horizontal="center")
        for i, (_, key, bucket) in enumerate(RULE_COLS, start=12):
            val = r.get(key, 0)
            cc = ws.cell(row=row, column=i, value="✓" if val else "")
            cc.alignment = Alignment(horizontal="center")
            cc.fill = GREEN if val else PatternFill()
        row += 1

    widths = [11, 24, 10, 12, 8, 8, 10, 10, 10, 8, 12] + [11] * len(RULE_COLS)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F5"

    # Rules tab
    ws2 = wb.create_sheet("Rules & Scoring")
    ws2["A1"] = "Rules & Scoring Logic"
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = HEAD
    ws2.merge_cells("A1:C1")
    rules_doc = [
        ("", "", ""),
        ("Indicator settings (match TradingView custom)", "", ""),
        ("BB %b",   "length=50, source=ohlc4, stdDev=2", ""),
        ("MACD",    "fast=3, slow=21, signal=9, source=close", ""),
        ("ADX",     "DI length=14 Wilder, ADX = SMA(14) of DX", ""),
        ("", "", ""),
        ("Scoring model", "Raw pts per group → normalized → group weighting", ""),
        ("BB group weight",   f"{GROUP_WEIGHT['BB']}% of composite", f"raw max {GROUP_MAX['BB']}"),
        ("MACD group weight", f"{GROUP_WEIGHT['MACD']}% of composite", f"raw max {GROUP_MAX['MACD']}"),
        ("ADX group weight",  f"{GROUP_WEIGHT['ADX']}% of composite", f"raw max {GROUP_MAX['ADX']}"),
        ("", "", ""),
        ("BB RULES", "Points", "Meaning"),
        ("BB LT Up",     f"{SIGNAL_WEIGHTS['BB_LT_uptrend']}",  "Close > 50-day MA"),
        ("BB 3d Up",     f"{SIGNAL_WEIGHTS['BB_ST_uptrend']}",  "BB% rising 3 days"),
        ("BB ≥ 0.8",     f"{SIGNAL_WEIGHTS['BB_val_strong']}",  "price near/above upper band (HIGHEST in BB)"),
        ("BB 0.1–0.3",   f"{SIGNAL_WEIGHTS['BB_bottom_zone']}", "bottom-zone bounce setup (mutex w/ ≥0.8)"),
        ("(Overbought)", "0",  "BB%≥1.3 — penalty removed per user"),
        ("", "", ""),
        ("MACD RULES", "Points", "Meaning"),
        ("MACD 3d Up",       f"{SIGNAL_WEIGHTS['MACD_3d_up']}",                     "MACD line rising 3 days"),
        ("Cross Above 0 (1–5d)", f"{SIGNAL_WEIGHTS['MACD_cross_above_zero_recent']}",   "MACD crossed Signal 1–5d ago, MACD ≥ 0 → STRONGEST"),
        ("Cross Near 0 (1–5d)",  f"{SIGNAL_WEIGHTS['MACD_cross_near_zero_recent']}",    "MACD crossed 1–5d ago, below 0 but close (|MACD|/60d<0.25)"),
        ("Cross Far Below (1–5d)",f"{SIGNAL_WEIGHTS['MACD_cross_far_below_recent']}",   "MACD crossed 1–5d ago, far below 0 → muted"),
        ("Cross 6–15d (sust.)",  f"{SIGNAL_WEIGHTS['MACD_cross_sustained']}",           "cross 6–15d ago, still holding — decaying signal"),
        ("MACD Pre-Cross",   f"{SIGNAL_WEIGHTS['MACD_closing_in']}",                "histogram rising, still <0 — pre-entry"),
        ("(Crossover >15d)", "0",  "crossovers older than 15d are neutral"),
        ("", "", ""),
        ("ADX RULES", "Points", "Meaning"),
        ("ADX Reversal",   f"{SIGNAL_WEIGHTS['ADX_reversal']}",    "rising 2–5d after prior decline (fresh trend)"),
        ("ADX Continuing", f"{SIGNAL_WEIGHTS['ADX_continuing']}",  "rising ≥6d (mutex w/ Reversal)"),
        ("ADX ≥ 18",       f"{SIGNAL_WEIGHTS['ADX_above_18']}",    "trend strong enough"),
        ("DI+ Rising",     f"{SIGNAL_WEIGHTS['DI+_rising']}",      "DI+ turning up"),
        ("DI+ Uptrend",    f"{SIGNAL_WEIGHTS['DI+_uptrend']}",     "DI+ up 3 days"),
        ("DI Gap Wide",    f"{SIGNAL_WEIGHTS['DI_gap_wide']}",     "DI+ − DI- ≥ 10"),
        ("DI Gap Widening",f"{SIGNAL_WEIGHTS['DI_gap_widening']}", "gap widening 2–5d"),
        ("", "", ""),
        ("Category", "Range", "Action"),
        ("Strong Buy", "≥ 75",  "high-conviction entry"),
        ("Watch",      "50–74", "monitor / partial"),
        ("Neutral",    "25–49", "weak / mixed"),
        ("Avoid",      "< 25",  "no alignment"),
    ]
    for i, row_data in enumerate(rules_doc, start=2):
        for j, v in enumerate(row_data, start=1):
            ws2.cell(row=i, column=j, value=v)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 30

    wb.save(path)


# ---------------------------------------------------------------------------
# HTML OUTPUT — with tooltips, add/remove, localStorage watchlist
# ---------------------------------------------------------------------------
def write_html(df: pd.DataFrame, path: Path, default_watchlist: list[str],
               breadth: dict | None = None) -> None:
    df_ok = df[df["status"] == "ok"].copy()

    def _sig_r(row, prefix=""):
        """Build the signal boolean dict, optionally prefixed (e.g. 'w_' for weekly)."""
        g = lambda k: int(row.get(prefix + k, 0) or 0)
        return {
            "BB_LT":  g("BB_LT_uptrend"),
            "BB_3d":  g("BB_ST_uptrend"),
            "BB_08":  g("BB_val_strong"),
            "BB_bot": g("BB_bottom_zone"),
            "M_3d":   g("MACD_3d_up"),
            "M_Xab":  g("MACD_cross_above_zero_recent"),
            "M_Xnr":  g("MACD_cross_near_zero_recent"),
            "M_Xfb":  g("MACD_cross_far_below_recent"),
            "M_Xsu":  g("MACD_cross_sustained"),
            "M_pre":  g("MACD_closing_in"),
            "A_rev":  g("ADX_reversal"),
            "A_con":  g("ADX_continuing"),
            "A_18":   g("ADX_above_18"),
            "DP_r":   g("DI+_rising"),
            "DP_u":   g("DI+_uptrend"),
            "G_w":    g("DI_gap_wide"),
            "G_wx":   g("DI_gap_widening"),
        }

    def _num(v, default=None):
        """Safe float conversion — returns default for NaN / None."""
        try:
            if v is None or pd.isna(v):
                return default
            return float(v)
        except Exception:
            return default

    payload = []
    for _, r in df_ok.iterrows():
        # Weekly payload (may be missing if <60 weekly bars)
        w_block = None
        if pd.notna(r.get("w_score")):
            w_block = {
                "cat":  r.get("w_category"),
                "sc":   _num(r.get("w_score"), 0),
                "bb":   _num(r.get("w_bb_pct"), 0),
                "m":    _num(r.get("w_macd"), 0),
                "ms":   _num(r.get("w_macd_sig"), 0),
                "mh":   _num(r.get("w_macd_hist"), 0),
                "a":    _num(r.get("w_adx"), 0),
                "dp":   _num(r.get("w_di_plus"), 0),
                "bs":   int(_num(r.get("w_bb_score"), 0)),
                "mSc":  int(_num(r.get("w_macd_score"), 0)),
                "aSc":  int(_num(r.get("w_adx_score"), 0)),
                "bR":   int(_num(r.get("w_bb_raw"), 0)),
                "mR":   int(_num(r.get("w_macd_raw"), 0)),
                "aR":   int(_num(r.get("w_adx_raw"), 0)),
                "dsc":  int(_num(r.get("w_dsc"), -1)),
                "r":    _sig_r(r, prefix="w_"),
            }

        payload.append({
            "t":    r["ticker"],
            "n":    r["name"],
            "ind":  r["industry"],
            "cl":   r["close"],
            "cat":  r["category"],
            "sc":   r["score"],
            "bb":   r["bb_pct"],
            "bbU":  r["bb_upper"],
            "bbL":  r["bb_lower"],
            "bbB":  r["bb_basis"],
            "m":    r["macd"],
            "ms":   r["macd_sig"],
            "mh":   r["macd_hist"],
            "a":    r["adx"],
            "dp":   r["di_plus"],
            "dm":   r["di_minus"],
            "ma":   r["ma50"],
            "bs":   r["bb_score"],
            "mSc":  r["macd_score"],
            "aSc":  r["adx_score"],
            "bR":   int(r["bb_raw"]),
            "mR":   int(r["macd_raw"]),
            "aR":   int(r["adx_raw"]),
            "dsc":  int(r["days_since_cross"]),
            "atr":  _num(r.get("atr"), 0),
            "stop": _num(r.get("stop"), 0),
            "stPct":_num(r.get("stop_pct"), 0),
            "r":    _sig_r(r),
            "w":    w_block,
            "dlt":  _num(r.get("delta")),         # prev-day score delta (None if no snapshot)
            "pCat": r.get("prev_category") or None,
        })
    payload_json   = json.dumps(payload, separators=(",", ":"), default=str)
    watchlist_json = json.dumps(default_watchlist)
    breadth_json   = json.dumps(breadth or {}, separators=(",", ":"), default=str)
    # Unique sorted sector list for the sector filter dropdown
    sectors = sorted({ (r["industry"] or "Unknown") for _, r in df_ok.iterrows() })
    sectors_json   = json.dumps(sectors)

    html = r"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Momentum Board</title>
<meta name="description" content="Nifty 500 momentum scanner — BB · MACD · ADX signals">
<meta property="og:title" content="Momentum Board">
<meta property="og:description" content="Nifty 500 momentum scanner — BB · MACD · ADX signals">
<meta property="og:type" content="website">
<meta name="twitter:card" content="summary">
<meta name="twitter:title" content="Momentum Board">
<meta name="twitter:description" content="Nifty 500 momentum scanner — BB · MACD · ADX signals">
<style>
  :root {
    color-scheme: dark;
    --bg:#0f1419; --panel:#1a1f2e; --ink:#e8eaed; --mute:#8b95a7;
    --green:#22c55e; --amber:#f59e0b; --red:#ef4444; --blue:#3b82f6;
    --line:#2a3142;
  }
  * { box-sizing:border-box }

  /* ----- Cross-browser dark scrollbars (belt-and-suspenders to color-scheme) ----- */
  html { scrollbar-color: #2a3142 #0f1419; scrollbar-width: thin }
  ::-webkit-scrollbar           { width:10px; height:10px }
  ::-webkit-scrollbar-track     { background:#0f1419 }
  ::-webkit-scrollbar-thumb     { background:#2a3142; border-radius:6px;
                                  border:2px solid #0f1419 }
  ::-webkit-scrollbar-thumb:hover { background:#3a4358 }
  ::-webkit-scrollbar-corner    { background:#0f1419 }

  /* ----- Force dark form controls across all browsers ----- */
  select, input, button, textarea {
    color:var(--ink); background-color:#0f1419;
    font-family:inherit;
  }
  select option {
    background:#1a1f2e; color:var(--ink);
  }
  select:focus, input:focus, textarea:focus { outline:none }
  html, body { height:100%; }
  body { margin:0; background:var(--bg); color:var(--ink);
    font:14px/1.5 -apple-system,Segoe UI,Inter,system-ui,sans-serif;
    display:flex; flex-direction:column; overflow:hidden; }
  .topbar, .breadth-row, .tabs-row, .controls, footer { flex-shrink:0 }
  .topbar {
    display:flex; align-items:center; gap:10px;
    padding:8px 32px; border-bottom:1px solid var(--line);
    background:linear-gradient(180deg,#1a1f2e,#0f1419);
    font-size:11px; color:var(--mute);
  }
  .topbar .brand {
    font-size:14px; font-weight:700; color:var(--ink);
    letter-spacing:-0.2px;
  }
  .topbar .brand-dot {
    display:inline-block; width:6px; height:6px; border-radius:50%;
    background:var(--blue); margin-right:8px; vertical-align:middle;
    box-shadow:0 0 8px rgba(59,130,246,0.6);
  }
  .topbar .sep { color:#3a4358; margin:0 4px }
  .topbar .dstamp { font-variant-numeric:tabular-nums; letter-spacing:0.2px }
  .topbar .spacer { flex:1 }
  .topbar .sync-state { font-size:10.5px; color:var(--mute); min-width:0; white-space:nowrap }
  .topbar .sync-state.ok   { color:#4ade80 }
  .topbar .sync-state.err  { color:#f87171 }
  .topbar .sync-state.busy { color:#fbbf24 }
  .topbar .icon-btn {
    background:transparent; border:1px solid var(--line); color:#cbd3e0;
    width:26px; height:22px; border-radius:5px; font-size:13px;
    cursor:pointer; display:inline-flex; align-items:center; justify-content:center;
  }
  .topbar .icon-btn:hover { border-color:var(--blue); color:var(--ink) }

  /* ---------- Sync dialog (modal) ---------- */
  .modal-bg {
    position:fixed; inset:0; background:rgba(0,0,0,0.55);
    display:none; align-items:center; justify-content:center; z-index:10000;
  }
  .modal-bg.open { display:flex }
  .modal {
    background:var(--panel); border:1px solid var(--line);
    border-radius:10px; padding:22px 24px; width:420px; max-width:92vw;
    box-shadow:0 20px 60px rgba(0,0,0,0.7);
  }
  .modal h3 { margin:0 0 6px; font-size:15px; color:var(--ink) }
  .modal .sub { color:var(--mute); font-size:11.5px; line-height:1.55; margin-bottom:14px }
  .modal label {
    display:block; font-size:11px; color:var(--mute);
    text-transform:uppercase; letter-spacing:0.4px; margin:10px 0 4px;
  }
  .modal input {
    width:100%; background:#0f1419; border:1px solid var(--line);
    color:var(--ink); padding:8px 10px; border-radius:6px;
    font:12px/1.4 'SF Mono',Menlo,Monaco,monospace; outline:none;
    box-sizing:border-box;
  }
  .modal input:focus { border-color:var(--blue) }
  .modal .row-btns {
    display:flex; gap:8px; margin-top:18px; align-items:center;
  }
  .modal .row-btns .spacer { flex:1 }
  .modal button {
    background:#1e3a5f; border:1px solid var(--line); color:var(--ink);
    padding:8px 14px; border-radius:7px; font-size:12px; cursor:pointer;
    font-family:inherit; font-weight:600;
  }
  .modal button:hover { border-color:var(--blue) }
  .modal button.primary { background:#2563eb; border-color:#2563eb }
  .modal button.danger  { background:#3a1e1e; border-color:#5b2a2a; color:#fca5a5 }
  .modal .msg { font-size:11px; color:var(--mute); margin-top:10px; min-height:14px }
  .modal .msg.ok   { color:#4ade80 }
  .modal .msg.err  { color:#f87171 }
  .modal .help {
    font-size:10.5px; color:var(--mute); line-height:1.6;
    background:#0b1017; border:1px dashed var(--line);
    border-radius:6px; padding:10px 12px; margin-top:14px;
  }
  .modal .help a { color:#60a5fa; text-decoration:none }
  .modal .help a:hover { text-decoration:underline }
  .modal .help code {
    background:#0f1419; padding:1px 5px; border-radius:3px;
    font:10.5px 'SF Mono',Menlo,monospace; color:#cbd3e0;
  }

  /* ---------- Breadth strip (market context bar) ---------- */
  .breadth-row {
    display:flex; gap:10px; padding:12px 32px;
    border-bottom:1px solid var(--line);
    background:#141926;
    overflow-x:auto;
  }
  .bcell {
    display:flex; flex-direction:column; gap:2px;
    min-width:110px; padding:4px 14px 4px 0;
    border-right:1px solid var(--line);
  }
  .bcell:last-child { border-right:none }
  .bcell .lbl { font-size:10px; color:var(--mute); text-transform:uppercase;
    letter-spacing:0.5px; font-weight:600 }
  .bcell .val { font-size:17px; font-weight:700; font-variant-numeric:tabular-nums }
  .bcell .sub2 { font-size:11px; color:var(--mute); font-variant-numeric:tabular-nums }
  .pos { color:#4ade80 } .neg { color:#fb7185 } .neu { color:var(--mute) }
  .sb-c { color:#4ade80 } .wa-c { color:#fbbf24 }
  .av-c { color:#fb7185 } .nu-c { color:var(--mute) }

  /* ---------- Tabs row ---------- */
  .tabs-row {
    display:flex; align-items:center; gap:8px;
    padding:12px 32px 0;
  }
  .tab {
    background:transparent; border:1px solid var(--line); color:var(--mute);
    padding:9px 18px; border-radius:8px 8px 0 0; font-size:13px;
    font-weight:600; cursor:pointer; font-family:inherit;
    border-bottom-color:transparent;
  }
  .tab:hover { color:#cbd3e0 }
  .tab.active {
    background:var(--panel); color:#e8eaed;
    border-color:var(--line) var(--line) var(--panel);
  }
  .tab .count { font-size:11px; color:var(--mute); margin-left:8px; font-weight:500 }
  .tab.active .count { color:#93c5fd }
  .tabs-row .spacer { flex:1 }
  /* Daily / Weekly toggle */
  .tf-toggle {
    display:inline-flex; background:#0f1419; border:1px solid var(--line);
    border-radius:8px; padding:3px; gap:0;
  }
  .tf-toggle button {
    background:transparent; border:none; color:var(--mute);
    padding:6px 14px; border-radius:6px; font-size:12px;
    font-weight:600; cursor:pointer; font-family:inherit;
    transition:background 0.15s;
  }
  .tf-toggle button.on {
    background:#1e3a5f; color:#e8eaed;
    box-shadow:inset 0 0 0 1px rgba(59,130,246,0.3);
  }
  .tf-toggle button:hover:not(.on) { color:#cbd3e0 }
  .controls { padding:10px 32px 12px; display:flex; gap:10px; flex-wrap:wrap;
    align-items:center; background:var(--panel);
    border-left:1px solid var(--line); border-right:1px solid var(--line);
    border-top:1px solid var(--line); }
  .controls input, .controls select, .controls button {
    background:#0f1419; border:1px solid var(--line); color:var(--ink);
    padding:7px 11px; border-radius:7px; font-size:12.5px; outline:none;
    font-family:inherit }
  .controls input:focus, .controls select:focus, .controls button:hover {
    border-color:var(--blue) }
  .controls button { cursor:pointer; background:#1e3a5f; font-weight:600 }
  .controls button.primary { background:#1e40af; border-color:#3b82f6 }
  .controls button.primary:hover { background:#2563eb }
  .controls .sep { color:var(--mute); font-size:12px }
  .controls .count { margin-left:auto }
  .wrap { padding:0 32px 12px; flex:1 1 auto; min-height:0;
    display:flex; flex-direction:column }

  /* ---------- Delta column / newly triggered badge / stop column ---------- */
  .delta { font-variant-numeric:tabular-nums; font-weight:600; text-align:right; font-size:12.5px }
  .delta.up  { color:#4ade80 }
  .delta.dn  { color:#fb7185 }
  .delta.zz  { color:var(--mute) }
  .new-badge {
    display:inline-block; margin-left:6px; padding:1px 5px;
    background:rgba(59,130,246,0.22); color:#93c5fd;
    border:1px solid rgba(59,130,246,0.5); border-radius:4px;
    font-size:9px; font-weight:700; letter-spacing:0.5px;
    vertical-align:middle;
  }
  .fresh-dot {
    display:inline-block; width:6px; height:6px; border-radius:50%;
    background:#3b82f6; box-shadow:0 0 6px rgba(59,130,246,0.6);
    margin-right:6px; vertical-align:middle;
  }
  .stop-cell {
    font-variant-numeric:tabular-nums; font-size:12px; color:#c3ccd9; text-align:right;
  }
  .stop-cell small { color:var(--mute); font-size:10.5px; display:block; margin-top:1px }
  /* ---- Two-table layout (NO sticky) ---- */
  .tbl-host {
    flex:1 1 auto; min-height:0;
    display:flex; flex-direction:column;
    background:var(--panel); border:1px solid var(--line);
    border-radius:10px; overflow:hidden;
  }
  .hdr-wrap {
    overflow:hidden;            /* no scrollbar on header */
    flex:0 0 auto;
    border-bottom:1px solid var(--line);
  }
  .body-wrap {
    overflow:auto;              /* ONLY scroll container */
    flex:1 1 auto;
    min-height:0;
  }
  table.ftbl {
    table-layout:fixed;
    border-collapse:separate; border-spacing:0;
    background:var(--panel);
    min-width:1560px;           /* enforce horizontal scroll below this */
    width:100%;
  }
  /* Scan view has a simpler column layout */
  table.ftbl.scan { min-width:1220px }
  th { color:#cbd3e0; font-weight:600; padding:10px 8px;
    text-align:left; font-size:11px; text-transform:uppercase;
    letter-spacing:0.3px; border-bottom:1px solid var(--line);
    white-space:nowrap }
  td { padding:10px 8px; border-bottom:1px solid var(--line); font-size:13px;
    vertical-align:middle; background:var(--panel);
    white-space:nowrap }
  tbody tr:hover td { background:#222a3d }

  /* Group header row (BB% / MACD / ADX) */
  thead tr.group-hd th {
    background:#1a1f2e; color:#cbd3e0;
    text-align:center; font-size:11px; font-weight:700;
    letter-spacing:0.5px;
    height:26px; padding:4px 8px;
  }
  thead tr.data-hd th { background:#232a3d }

  /* ---- Group header cells: uniform dark bg, subtle top accent line ---- */
  thead tr.group-hd th:nth-child(2),
  thead tr.group-hd th:nth-child(3),
  thead tr.group-hd th:nth-child(4) {
    background:#1c2233; color:#9ba5b8;
    box-shadow:inset 0 2px 0 0 rgba(148,163,184,0.18);
  }

  /* ---- Vertical group dividers (watchlist view) ----
     Col  8 = BB start  (6 base + 1 + 1)
     Col 11 = MACD start (8 + 3)
     Col 16 = ADX start  (11 + 5)
     Col 19 = Stop       (16 + 3)
     Col 20 = Remove     (19 + 1) */
  thead tr.data-hd th:nth-child(8),
  thead tr.data-hd th:nth-child(11),
  thead tr.data-hd th:nth-child(16),
  tbody td:nth-child(8),
  tbody td:nth-child(11),
  tbody td:nth-child(16) {
    border-left:1px solid #323a4f;
    box-shadow:inset 1px 0 0 rgba(255,255,255,0.025);
  }
  thead tr.data-hd th:nth-child(19),
  tbody td:nth-child(19),
  thead tr.data-hd th:nth-child(20),
  tbody td:nth-child(20) { border-left:1px solid var(--line) }

  /* Sticky first column (left pin) inside each of the two tables */
  thead tr.data-hd th:first-child,
  thead tr.group-hd th:first-child,
  tbody td:first-child {
    position:sticky; left:0; z-index:2;
    box-shadow:2px 0 4px rgba(0,0,0,0.25);
  }
  thead tr.data-hd th:first-child { background:#232a3d }
  thead tr.group-hd th:first-child { background:#1a1f2e }
  tbody td:first-child { background:#1e2333 }
  tbody tr:hover td:first-child { background:#262e44 }
  .tk { font-weight:700; color:#fff }
  .tk small { display:block; color:var(--mute); font-weight:400; font-size:11px;
    margin-top:2px; white-space:normal; line-height:1.3 }
  .num { font-variant-numeric:tabular-nums; color:#c3ccd9 }
  .score { font-variant-numeric:tabular-nums; font-size:15px; font-weight:700 }
  .sig { text-align:center; font-weight:700; cursor:help; position:relative }
  .sig.on { background:rgba(34,197,94,0.15); color:#4ade80 }
  .sig.off { color:#3a4256 }
  .sig.pen { background:rgba(239,68,68,0.2); color:#f87171 }
  .chip { display:inline-block; padding:3px 10px; border-radius:999px;
    font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:0.3px }
  .chip.sb { background:rgba(34,197,94,0.18); color:#4ade80;
    border:1px solid rgba(34,197,94,0.4) }
  .chip.wa { background:rgba(245,158,11,0.18); color:#fbbf24;
    border:1px solid rgba(245,158,11,0.4) }
  .chip.nu { background:rgba(139,149,167,0.18); color:#cbd3e0;
    border:1px solid rgba(139,149,167,0.4) }
  .chip.av { background:rgba(239,68,68,0.18); color:#fb7185;
    border:1px solid rgba(239,68,68,0.4) }
  .rm { background:transparent; border:1px solid var(--line); color:var(--mute);
    padding:3px 8px; font-size:11px; border-radius:6px; cursor:pointer }
  .rm:hover { background:rgba(239,68,68,0.12); color:#fb7185;
    border-color:#fb7185 }
  .group-hd { background:#1a1f2e; color:var(--mute); text-align:center;
    font-size:10px; border-bottom:1px solid var(--line) }
  /* Floating tooltip (JS-driven so it escapes any scroll container) */
  #floatTip {
    position:fixed; background:#0b0f17; color:#e8eaed;
    padding:8px 12px; border-radius:6px; font-size:11px;
    font-family:'SF Mono',Menlo,Monaco,monospace; line-height:1.65;
    border:1px solid #3b82f6; z-index:9999;
    box-shadow:0 8px 20px rgba(0,0,0,0.5);
    pointer-events:none; white-space:pre; display:none; max-width:420px;
  }
  [data-tip] { cursor:help }
  footer { padding:16px 32px; color:var(--mute); font-size:12px }
  .saved-tag { font-size:11px; color:var(--green); margin-left:8px }
  .count { font-size:12px; color:var(--mute); margin-left:auto }

  /* ----- Settings / logic panel at the bottom ----- */
  .panel-wrap { padding:12px 32px 24px; flex-shrink:0 }
  details.settings {
    background:var(--panel); border:1px solid var(--line);
    border-radius:10px;
  }
  details.settings > summary {
    list-style:none; cursor:pointer; padding:14px 18px;
    font-size:13px; font-weight:600; color:#cbd3e0;
    display:flex; align-items:center; gap:10px;
  }
  details.settings > summary::-webkit-details-marker { display:none }
  details.settings > summary::before {
    content:"▸"; color:var(--blue); font-size:12px;
    transition:transform 0.15s;
  }
  details.settings[open] > summary::before { transform:rotate(90deg) }
  details.settings > summary .spacer { flex:1 }
  details.settings > summary .hint { font-weight:400; color:var(--mute); font-size:11px }

  .settings-body {
    padding:6px 22px 22px; border-top:1px solid var(--line);
    max-height:45vh; overflow-y:auto;
  }
  .settings-note {
    font-size:12px; color:var(--mute); margin:12px 0 18px;
    line-height:1.6;
  }
  .settings-note b { color:#cbd3e0 }
  .wgroups {
    display:grid; grid-template-columns:repeat(auto-fit,minmax(250px,1fr));
    gap:22px 28px;
  }
  .wcol h3 {
    margin:0 0 10px; font-size:12px; text-transform:uppercase;
    letter-spacing:0.4px; border-bottom:1px solid var(--line);
    padding-bottom:6px;
  }
  .wcol label {
    display:flex; align-items:center; justify-content:space-between;
    gap:12px; padding:4px 0; font-size:12px; color:#cbd3e0;
  }
  .wcol label .lbl { flex:1; color:#cbd3e0 }
  .wcol label .mute { color:var(--mute); font-size:11px; margin-left:6px }
  .wcol input[type="number"] {
    width:56px; background:#0f1419; border:1px solid var(--line);
    color:#e8eaed; padding:4px 6px; border-radius:5px; font-size:12px;
    text-align:right; font-variant-numeric:tabular-nums; outline:none;
  }
  .wcol input[type="number"]:focus { border-color:var(--blue) }
  .settings-actions {
    margin-top:20px; padding-top:14px; border-top:1px solid var(--line);
    display:flex; gap:10px; align-items:center;
  }
  .settings-actions button {
    background:#1e3a5f; border:1px solid var(--line); color:var(--ink);
    padding:7px 14px; border-radius:7px; font-size:12px; cursor:pointer;
    font-family:inherit; font-weight:600;
  }
  .settings-actions button:hover { border-color:var(--blue) }
  .settings-actions .status { font-size:11px; color:var(--mute) }
  .settings-actions .status b { color:#4ade80 }
</style></head>
<body>
<div class="topbar">
  <span class="brand"><span class="brand-dot"></span>Momentum Board</span>
  <span class="sep">·</span>
  <span class="dstamp">__STAMP_HUMAN__</span>
  <span class="spacer"></span>
  <span id="syncState" class="sync-state"></span>
  <button class="icon-btn" onclick="openSyncDialog()" title="Cross-device sync settings">⇄</button>
</div>

<!-- Sync dialog -->
<div class="modal-bg" id="syncModal" onclick="if(event.target===this)closeSyncDialog()">
  <div class="modal">
    <h3>Cross-device watchlist sync</h3>
    <div class="sub">Syncs your watchlist between phone, tablet, and desktop via a private GitHub Gist. Paste the same Gist ID + token on every device.</div>
    <label>Gist ID</label>
    <input id="syncGistId" placeholder="e.g. 3f2a5b9c0d8e1a2b3c4d">
    <label>Personal access token (gist scope)</label>
    <input id="syncToken" placeholder="ghp_... or github_pat_...">
    <div class="msg" id="syncMsg"></div>
    <div class="help">
      <b>One-time setup:</b><br>
      1. Create a secret gist at <a href="https://gist.github.com" target="_blank">gist.github.com</a> with filename <code>watchlist.json</code> and body <code>[]</code>. Copy the long ID from its URL.<br>
      2. Create a classic token at <a href="https://github.com/settings/tokens/new?scopes=gist&description=watchlist-sync" target="_blank">github.com/settings/tokens</a> — check only <code>gist</code> scope. Copy the token (shown once).<br>
      3. Paste both above and hit Save. Repeat on every device with the same values.
    </div>
    <div class="row-btns">
      <button class="danger" onclick="disableSync()" id="syncDisableBtn">Disable sync</button>
      <span class="spacer"></span>
      <button onclick="closeSyncDialog()">Cancel</button>
      <button class="primary" onclick="saveSyncConfig()">Save & test</button>
    </div>
  </div>
</div>

<!-- Market breadth strip -->
<div class="breadth-row" id="breadthRow"></div>

<!-- Tabs + daily/weekly toggle -->
<div class="tabs-row">
  <button class="tab active" id="tabWL" onclick="setTab('wl')">
    My Watchlist <span class="count" id="tabWLCount"></span>
  </button>
  <button class="tab" id="tabSB" onclick="setTab('sb')">
    Strong Buy Scan <span class="count" id="tabSBCount"></span>
  </button>
  <span class="spacer"></span>
  <div class="tf-toggle">
    <button class="on" id="tfD" onclick="setTf('d')">Daily</button>
    <button id="tfW" onclick="setTf('w')">Weekly</button>
  </div>
</div>

<!-- Controls bar -->
<div class="controls">
  <input id="q" placeholder="Search…" oninput="render()" style="min-width:160px">
  <select id="cat" onchange="render()">
    <option value="">All categories</option>
    <option>Strong Buy</option><option>Watch</option>
    <option>Neutral</option><option>Avoid</option>
  </select>
  <select id="sector" onchange="render()"></select>
  <select id="sort" onchange="render()">
    <option value="score-desc">Sort: Score ↓</option>
    <option value="score-asc">Sort: Score ↑</option>
    <option value="delta-desc">Sort: Δ ↓</option>
    <option value="ticker">Sort: Ticker A–Z</option>
  </select>
  <span class="sep" id="wlOnly">|</span>
  <select id="addDD" style="min-width:240px"></select>
  <button class="primary" onclick="addTicker()">+ Add to watchlist</button>
  <button onclick="resetWatchlist()">Reset WL</button>
  <span class="sep" id="sbOnly" style="display:none">|</span>
  <label id="sbScoreWrap" style="display:none; font-size:11px; color:var(--mute); gap:6px; align-items:center">
    Min score <input id="sbMin" type="number" value="70" step="5" min="0" max="100" style="width:56px; padding:4px 6px">
  </label>
  <span id="savedTag" class="saved-tag"></span>
  <span class="count" id="countTag"></span>
</div>

<div class="wrap">
  <div class="tbl-host">
    <div class="hdr-wrap" id="hdrWrap">
      <table class="ftbl" id="hdrTbl">
        <colgroup id="cgHdr"></colgroup>
        <thead id="thWrap"></thead>
      </table>
    </div>
    <div class="body-wrap" id="bodyWrap">
      <table class="ftbl" id="bodyTbl">
        <colgroup id="cgBody"></colgroup>
        <tbody id="tb"></tbody>
      </table>
    </div>
  </div>
</div>

<div class="panel-wrap">
  <details class="settings" id="settingsPanel">
    <summary>
      Scoring logic &amp; weights
      <span class="spacer"></span>
      <span class="hint">click to expand / edit · persists in your browser</span>
    </summary>
    <div class="settings-body">
      <div class="settings-note">
        Composite =
        <b>(BB_raw / BB_max)</b> &times; BB%  +
        <b>(MACD_raw / MACD_max)</b> &times; MACD%  +
        <b>(ADX_raw / ADX_max)</b> &times; ADX%.<br>
        Each group's raw points sum the weights of whichever signals are active for a ticker.
        Group max auto-updates to the best achievable combination (honouring mutex rules:
        BB ≥0.8 vs 0.1–0.3, MACD cross buckets mutually exclusive, ADX Reversal vs Continuing).
        <b>Hidden columns</b> (0.1–0.3, Pre, Rev, Cont, Gap, DI+3d) are still scored in the background —
        they just aren't shown in the grid.
      </div>

      <div class="wgroups">
        <div class="wcol">
          <h3 style="color:#cbd3e0">Group weights (%)</h3>
          <label><span class="lbl">BB %b</span>  <input type="number" data-gw="BB"   step="5" min="0" max="100"></label>
          <label><span class="lbl">MACD</span>   <input type="number" data-gw="MACD" step="5" min="0" max="100"></label>
          <label><span class="lbl">ADX</span>    <input type="number" data-gw="ADX"  step="5" min="0" max="100"></label>
          <label style="margin-top:8px"><span class="mute" id="gwSum">Total: 100</span></label>
        </div>

        <div class="wcol">
          <h3 style="color:#9ba5b8">BB signals</h3>
          <label><span class="lbl">LT Up <span class="mute">(close &gt; 50MA)</span></span>     <input type="number" data-sw="BB_LT_uptrend"></label>
          <label><span class="lbl">3d Up <span class="mute">(BB% rising 3d)</span></span>       <input type="number" data-sw="BB_ST_uptrend"></label>
          <label><span class="lbl">≥ 0.8 <span class="mute">(strong) — VISIBLE</span></span>    <input type="number" data-sw="BB_val_strong"></label>
          <label><span class="lbl">0.1–0.3 <span class="mute">(bottom) — hidden</span></span>  <input type="number" data-sw="BB_bottom_zone"></label>
        </div>

        <div class="wcol">
          <h3 style="color:#9ba5b8">MACD signals</h3>
          <label><span class="lbl">3d Up <span class="mute">(line rising 3d)</span></span>                     <input type="number" data-sw="MACD_3d_up"></label>
          <label><span class="lbl">Cross Above 0 <span class="mute">(1-5d, ≥ 0) STRONGEST</span></span>        <input type="number" data-sw="MACD_cross_above_zero_recent"></label>
          <label><span class="lbl">Cross Near 0 <span class="mute">(1-5d, close to 0)</span></span>           <input type="number" data-sw="MACD_cross_near_zero_recent"></label>
          <label><span class="lbl">Cross Far Below <span class="mute">(1-5d, muted)</span></span>            <input type="number" data-sw="MACD_cross_far_below_recent"></label>
          <label><span class="lbl">Cross 6-15d <span class="mute">(sustained)</span></span>                  <input type="number" data-sw="MACD_cross_sustained"></label>
          <label><span class="lbl">Pre-cross <span class="mute">(hist↑, &lt;0) — hidden</span></span>       <input type="number" data-sw="MACD_closing_in"></label>
        </div>

        <div class="wcol">
          <h3 style="color:#9ba5b8">ADX signals</h3>
          <label><span class="lbl">Reversal <span class="mute">(2-5d up) — hidden</span></span>    <input type="number" data-sw="ADX_reversal"></label>
          <label><span class="lbl">Continuing <span class="mute">(≥6d up) — hidden</span></span>   <input type="number" data-sw="ADX_continuing"></label>
          <label><span class="lbl">≥ 18 <span class="mute">— VISIBLE</span></span>                  <input type="number" data-sw="ADX_above_18"></label>
          <label><span class="lbl">DI+ Rising <span class="mute">— VISIBLE</span></span>            <input type="number" data-sw="DI+_rising"></label>
          <label><span class="lbl">DI+ 3d Up <span class="mute">— hidden</span></span>              <input type="number" data-sw="DI+_uptrend"></label>
          <label><span class="lbl">Gap Wide <span class="mute">(≥10) — hidden</span></span>         <input type="number" data-sw="DI_gap_wide"></label>
          <label><span class="lbl">Gap Widening <span class="mute">— VISIBLE</span></span>          <input type="number" data-sw="DI_gap_widening"></label>
        </div>
      </div>

      <div class="settings-actions">
        <button onclick="resetWeights()">Reset to defaults</button>
        <span class="status" id="wStatus"></span>
      </div>
    </div>
  </details>
</div>

<footer>
  Data: Yahoo Finance · Indicators computed locally to match TradingView custom settings ·
  Watchlist &amp; weights persist in your browser (localStorage) · Educational use only.
</footer>

<script>
const DATA      = __DATA__;
const DEFAULT_WL= __WATCHLIST__;
const BREADTH   = __BREADTH__;
const SECTORS   = __SECTORS__;
const STAMP         = "__STAMP__";
const LS_KEY        = "am_watch_v2";
const LS_W_KEY      = "am_watch_v2_weights";
const LS_GW_KEY     = "am_watch_v2_gweights";
const LS_TAB        = "am_watch_v2_tab";
const LS_TF         = "am_watch_v2_tf";
const LS_SYNC_GIST  = "am_watch_v2_sync_gist";
const LS_SYNC_TOKEN = "am_watch_v2_sync_token";
const LS_SYNC_TS    = "am_watch_v2_sync_ts";
// -------- View state --------
let TAB = localStorage.getItem(LS_TAB) || "wl";   // wl | sb
let TF  = localStorage.getItem(LS_TF)  || "d";    // d | w

// ------------------------------------------------------------------
// Scoring config (editable at runtime via settings panel)
// ------------------------------------------------------------------
const DEFAULT_WEIGHTS = {
  // BB group
  BB_LT_uptrend:  6,
  BB_ST_uptrend:  8,
  BB_val_strong: 12,
  BB_bottom_zone: 8,
  // MACD group (all cross buckets are mutually exclusive)
  MACD_3d_up:                    4,
  MACD_cross_above_zero_recent: 20,
  MACD_cross_near_zero_recent:  14,
  MACD_cross_far_below_recent:   4,
  MACD_cross_sustained:          2,
  MACD_closing_in:               4,
  // ADX group
  ADX_reversal:  12,
  ADX_continuing: 6,
  ADX_above_18:   5,
  "DI+_rising":   8,
  "DI+_uptrend":  8,
  DI_gap_wide:    5,
  DI_gap_widening:12,
};
const DEFAULT_GW = {BB: 30, MACD: 40, ADX: 30};

// Short-key → full signal name (payload.r uses short keys for size)
const SIG_MAP = {
  BB_LT: "BB_LT_uptrend",  BB_3d: "BB_ST_uptrend",
  BB_08: "BB_val_strong",  BB_bot:"BB_bottom_zone",
  M_3d:  "MACD_3d_up",
  M_Xab: "MACD_cross_above_zero_recent",
  M_Xnr: "MACD_cross_near_zero_recent",
  M_Xfb: "MACD_cross_far_below_recent",
  M_Xsu: "MACD_cross_sustained",
  M_pre: "MACD_closing_in",
  A_rev: "ADX_reversal",    A_con: "ADX_continuing",
  A_18:  "ADX_above_18",
  DP_r:  "DI+_rising",      DP_u:  "DI+_uptrend",
  G_w:   "DI_gap_wide",     G_wx:  "DI_gap_widening",
};
const GROUP_OF = {
  BB_LT_uptrend:"BB", BB_ST_uptrend:"BB", BB_val_strong:"BB", BB_bottom_zone:"BB",
  MACD_3d_up:"MACD", MACD_cross_above_zero_recent:"MACD",
  MACD_cross_near_zero_recent:"MACD", MACD_cross_far_below_recent:"MACD",
  MACD_cross_sustained:"MACD", MACD_closing_in:"MACD",
  ADX_reversal:"ADX", ADX_continuing:"ADX", ADX_above_18:"ADX",
  "DI+_rising":"ADX", "DI+_uptrend":"ADX",
  DI_gap_wide:"ADX", DI_gap_widening:"ADX",
};

let WEIGHTS, GW;
function loadSettings() {
  try {
    const s = localStorage.getItem(LS_W_KEY);
    WEIGHTS = s ? {...DEFAULT_WEIGHTS, ...JSON.parse(s)} : {...DEFAULT_WEIGHTS};
  } catch(e) { WEIGHTS = {...DEFAULT_WEIGHTS}; }
  try {
    const s = localStorage.getItem(LS_GW_KEY);
    GW = s ? {...DEFAULT_GW, ...JSON.parse(s)} : {...DEFAULT_GW};
  } catch(e) { GW = {...DEFAULT_GW}; }
}
function saveSettings() {
  localStorage.setItem(LS_W_KEY, JSON.stringify(WEIGHTS));
  localStorage.setItem(LS_GW_KEY, JSON.stringify(GW));
  flashSaved();
}
loadSettings();

// Helper used by rowHtml tooltips to show current weight
function W(key) { return WEIGHTS[key] ?? 0; }

// Compute max achievable per group given current weights (honours mutex rules)
function groupMax() {
  const bb = WEIGHTS.BB_LT_uptrend + WEIGHTS.BB_ST_uptrend
           + Math.max(WEIGHTS.BB_val_strong, WEIGHTS.BB_bottom_zone);
  const macd = WEIGHTS.MACD_3d_up + Math.max(
    WEIGHTS.MACD_cross_above_zero_recent,
    WEIGHTS.MACD_cross_near_zero_recent,
    WEIGHTS.MACD_cross_far_below_recent,
    WEIGHTS.MACD_cross_sustained,
    WEIGHTS.MACD_closing_in
  );
  const adx = Math.max(WEIGHTS.ADX_reversal, WEIGHTS.ADX_continuing)
            + WEIGHTS.ADX_above_18
            + WEIGHTS["DI+_rising"] + WEIGHTS["DI+_uptrend"]
            + WEIGHTS.DI_gap_wide + WEIGHTS.DI_gap_widening;
  return {BB: bb || 1, MACD: macd || 1, ADX: adx || 1};
}

// Rescore a single record in place. Applies to both daily (top-level) and
// weekly (nested under d.w) if present.
function _scoreObj(obj, GM) {
  let bbRaw=0, mRaw=0, aRaw=0;
  for (const [k, full] of Object.entries(SIG_MAP)) {
    if (!obj.r || !obj.r[k]) continue;
    const w = WEIGHTS[full] || 0;
    const g = GROUP_OF[full];
    if      (g === "BB")   bbRaw  += w;
    else if (g === "MACD") mRaw   += w;
    else if (g === "ADX")  aRaw   += w;
  }
  const bbP = Math.min(1, Math.max(0, bbRaw / GM.BB));
  const mP  = Math.min(1, Math.max(0, mRaw  / GM.MACD));
  const aP  = Math.min(1, Math.max(0, aRaw  / GM.ADX));
  const sc  = bbP * GW.BB + mP * GW.MACD + aP * GW.ADX;
  let cat;
  if      (sc >= 75) cat = "Strong Buy";
  else if (sc >= 50) cat = "Watch";
  else if (sc >= 25) cat = "Neutral";
  else               cat = "Avoid";
  obj.bR = bbRaw; obj.mR = mRaw; obj.aR = aRaw;
  obj.bs = Math.round(bbP*100); obj.mSc = Math.round(mP*100); obj.aSc = Math.round(aP*100);
  obj.sc = Math.round(sc*10)/10; obj.cat = cat;
  obj._gm = GM;
}
function computeScore(d) {
  const GM = groupMax();
  _scoreObj(d, GM);            // daily (top level)
  if (d.w) _scoreObj(d.w, GM); // weekly (nested)
}
function rescoreAll() { DATA.forEach(computeScore); }
rescoreAll();

// Accessor that reads either the daily or the weekly view of a row.
// Returns the row itself for daily, or d.w for weekly (with fallback to daily).
function V(d) { return (TF === "w" && d.w) ? d.w : d; }

// Sync horizontal scroll of header table with body table
(function(){
  const body = document.getElementById('bodyWrap');
  const hdr  = document.getElementById('hdrWrap');
  if (body && hdr) {
    body.addEventListener('scroll', () => { hdr.scrollLeft = body.scrollLeft; });
  }
})();

// Floating tooltip (attached to <body> so it escapes any scroll clipping)
(function(){
  const tip = document.createElement('div');
  tip.id = 'floatTip';
  document.body.appendChild(tip);
  let cur = null;
  document.addEventListener('mouseover', e => {
    const el = e.target.closest('[data-tip]');
    if (!el || el === cur) return;
    cur = el;
    tip.textContent = el.getAttribute('data-tip');
    tip.style.display = 'block';
  });
  document.addEventListener('mousemove', e => {
    if (!cur) return;
    const pad = 14;
    const r = tip.getBoundingClientRect();
    let x = e.clientX + pad;
    let y = e.clientY + pad;
    if (x + r.width > window.innerWidth - 6)  x = e.clientX - r.width - pad;
    if (y + r.height > window.innerHeight - 6) y = e.clientY - r.height - pad;
    if (x < 4) x = 4;
    if (y < 4) y = 4;
    tip.style.left = x + 'px';
    tip.style.top  = y + 'px';
  });
  document.addEventListener('mouseout', e => {
    const el = e.target.closest('[data-tip]');
    if (el === cur) { tip.style.display = 'none'; cur = null; }
  });
})();

// Build lookup
const BY_T = {};
DATA.forEach(d => BY_T[d.t] = d);

function loadWL() {
  try {
    const s = localStorage.getItem(LS_KEY);
    if (s) return JSON.parse(s);
  } catch (e) {}
  return [...DEFAULT_WL];
}
function saveWL(wl) {
  localStorage.setItem(LS_KEY, JSON.stringify(wl));
  flashSaved();
  // Push to sync (debounced) if configured
  if (typeof syncPushDebounced === "function") syncPushDebounced();
}
function flashSaved() {
  const t = document.getElementById("savedTag");
  t.textContent = "✓ saved";
  setTimeout(() => t.textContent = "", 1400);
}

let watchlist = loadWL();

function catClass(c) {
  return {"Strong Buy":"sb","Watch":"wa","Neutral":"nu","Avoid":"av"}[c] || "nu";
}
function sig(on, tip) {
  if (on === undefined || on === null) return `<td class="sig off"></td>`;
  return `<td class="sig ${on ? 'on' : 'off'}" data-tip="${tip}">${on ? '✓' : ''}</td>`;
}

// Delta cell — receives a numeric delta or null
function deltaCell(dlt) {
  if (dlt === null || dlt === undefined) return `<td class="delta zz">–</td>`;
  const v = Number(dlt);
  if (Math.abs(v) < 0.05) return `<td class="delta zz">0.0</td>`;
  const arrow = v > 0 ? "▲" : "▼";
  const cls = v > 0 ? "up" : "dn";
  return `<td class="delta ${cls}" data-tip="Δ vs previous session snapshot">${arrow} ${Math.abs(v).toFixed(1)}</td>`;
}

// Category chip with NEW / FRESH markers
function catChip(d, V) {
  const base = `<span class="chip ${catClass(V.cat)}" data-tip="Thresholds:&#10;≥75 Strong Buy · 50–74 Watch · 25–49 Neutral · <25 Avoid">${V.cat}</span>`;
  // NEW = category promoted into Strong Buy today (vs. prev snapshot — daily mode only)
  const newBadge = (TF === "d" && V.cat === "Strong Buy"
    && d.pCat && d.pCat !== "Strong Buy")
    ? `<span class="new-badge" data-tip="Newly triggered Strong Buy — was '${d.pCat}' yesterday">NEW</span>` : "";
  // FRESH dot = MACD crossover happened today or yesterday
  const fresh = (V.dsc !== undefined && V.dsc >= 0 && V.dsc <= 1)
    ? `<span class="fresh-dot" data-tip="MACD crossover ${V.dsc} day(s) ago"></span>` : "";
  return `<td>${fresh}${base}${newBadge}</td>`;
}

function stopCell(d) {
  if (!d.stop || d.stop <= 0) return `<td class="stop-cell"><span class="neu">–</span></td>`;
  return `<td class="stop-cell" data-tip="ATR-based stop: close − 2×ATR(14)&#10;ATR = ${d.atr}&#10;Stop = ${d.stop.toFixed(2)}&#10;Risk = ${d.stPct.toFixed(2)}%">${d.stop.toFixed(2)}<small>${d.stPct.toFixed(1)}%</small></td>`;
}

// Full 20-col watchlist row
function rowHtmlWL(d) {
  const v = V(d);
  const macdDirection = v.mh > 0 ? "↑ bullish" : "↓ bearish";
  const diGap = (v.dp - (v.dm ?? 0)).toFixed(2);
  const adxTip = `ADX   = ${v.a}\nDI+   = ${v.dp}\nDI−   = ${v.dm ?? "-"}\nGap   = ${diGap}`;
  const dscStr = v.dsc < 0 ? "no recent cross" : `${v.dsc} day(s) ago`;
  const GM = v._gm || groupMax();
  const scoreTip =
    `BB   ${v.bR}/${GM.BB} → ${v.bs}%\n` +
    `MACD ${v.mR}/${GM.MACD} → ${v.mSc}%\n` +
    `ADX  ${v.aR}/${GM.ADX} → ${v.aSc}%\n` +
    `Composite: ${v.sc}\n` +
    `Group weights: BB ${GW.BB}% · MACD ${GW.MACD}% · ADX ${GW.ADX}%` +
    (TF === "w" ? "\n[WEEKLY timeframe]" : "");
  const closeTip = `Close     = ${d.cl}\n50-day MA = ${d.ma}\nIndustry  = ${d.ind}`;
  const bbTip =
    `BB %b = ${(v.bb ?? 0).toFixed(3)}\n` +
    (TF === "d" ? `Upper = ${d.bbU}\nBasis = ${d.bbB}\nLower = ${d.bbL}` : "");

  return `<tr data-cat="${v.cat}" data-score="${v.sc}" data-t="${d.t}">
    <td class="tk" data-tip="${d.n.replace(/"/g,'&quot;')}">${d.t}<small>${d.ind}</small></td>
    <td class="num" data-tip="${closeTip}">${d.cl.toFixed(2)}</td>
    ${deltaCell(d.dlt)}
    ${catChip(d, v)}
    <td class="score" data-tip="${scoreTip}">${v.sc.toFixed(1)}</td>
    <td class="num" data-tip="${bbTip}">${(v.bb ?? 0).toFixed(3)}</td>
    <td class="num" data-tip="${adxTip}">${(v.a ?? 0).toFixed(1)}</td>
    ${sig(v.r.BB_LT,  `Close > 50-day MA  (${W('BB_LT_uptrend')} pts)`)}
    ${sig(v.r.BB_3d,  `BB% rising 3 days  (${W('BB_ST_uptrend')} pts)`)}
    ${sig(v.r.BB_08,  `BB% ≥ 0.8 — strongest BB  (${W('BB_val_strong')} pts)`)}
    ${sig(v.r.M_3d,   `MACD 3-day up  (${W('MACD_3d_up')} pts)`)}
    ${sig(v.r.M_Xab,  `MACD crossed Signal 1–5d ago, ≥0 → STRONGEST  (${W('MACD_cross_above_zero_recent')} pts) · ${dscStr}`)}
    ${sig(v.r.M_Xnr,  `MACD crossed 1–5d ago, below 0 but close  (${W('MACD_cross_near_zero_recent')} pts) · ${dscStr}`)}
    ${sig(v.r.M_Xfb,  `MACD crossed 1–5d ago, far below 0 — muted  (${W('MACD_cross_far_below_recent')} pts) · ${dscStr}`)}
    ${sig(v.r.M_Xsu,  `MACD crossover 6–15d ago — sustained  (${W('MACD_cross_sustained')} pts) · ${dscStr}`)}
    ${sig(v.r.A_18,   `ADX ≥ 18  (${W('ADX_above_18')} pts)`)}
    ${sig(v.r.DP_r,   `DI+ rising  (${W('DI+_rising')} pts)`)}
    ${sig(v.r.G_wx,   `DI gap widening ≥2 consecutive days  (${W('DI_gap_widening')} pts)`)}
    ${stopCell(d)}
    <td><button class="rm" onclick="removeTicker('${d.t}')">Remove</button></td>
  </tr>`;
}

// Simpler 10-col scan row (Strong Buy scan of Nifty 500)
function rowHtmlSB(d) {
  const v = V(d);
  const dscStr = v.dsc < 0 ? "no recent cross" : `${v.dsc} day(s) ago`;
  const fresh = (v.dsc !== undefined && v.dsc >= 0 && v.dsc <= 1)
    ? `<span class="fresh-dot" data-tip="Fresh MACD crossover ${v.dsc}d ago"></span>` : "";
  const newBadge = (TF === "d" && v.cat === "Strong Buy"
    && d.pCat && d.pCat !== "Strong Buy")
    ? `<span class="new-badge">NEW</span>` : "";
  const catTip = "Thresholds: ≥75 Strong Buy · 50–74 Watch · 25–49 Neutral · <25 Avoid";
  return `<tr data-t="${d.t}">
    <td class="tk" data-tip="${d.n.replace(/"/g,'&quot;')}">${fresh}${d.t}<small>${d.ind}</small></td>
    <td class="num">${d.cl.toFixed(2)}</td>
    ${deltaCell(d.dlt)}
    <td><span class="chip ${catClass(v.cat)}" data-tip="${catTip}">${v.cat}</span>${newBadge}</td>
    <td class="score">${v.sc.toFixed(1)}</td>
    <td class="num">${(v.bb ?? 0).toFixed(3)}</td>
    <td class="num">${(v.a ?? 0).toFixed(1)}</td>
    <td class="num" data-tip="Days since MACD crossover">${v.dsc < 0 ? '–' : v.dsc + 'd'}</td>
    ${stopCell(d)}
    <td><button class="rm" style="background:#1e3a5f;color:#93c5fd;border-color:#3b82f6" onclick="addTickerDirect('${d.t}')">+ Add</button></td>
  </tr>`;
}

// -------- Dynamic colgroup / thead generation per tab --------
const COLS_WL = {
  widths: [170,86,62,110,70,68,60, 58,58,62, 60,72,72,72,82, 58,62,68, 96,82],
  thead: `
    <tr class="group-hd">
      <th colspan="7"></th>
      <th colspan="3">BB% · <span data-gwlbl="BB">30</span>%</th>
      <th colspan="5">MACD · <span data-gwlbl="MACD">40</span>%</th>
      <th colspan="3">ADX · <span data-gwlbl="ADX">30</span>%</th>
      <th colspan="2"></th>
    </tr>
    <tr class="data-hd">
      <th>Ticker</th><th>Close</th><th data-tip="Day-over-day Δ in composite score">Δ</th><th>Category</th><th>Score</th>
      <th>BB%</th><th>ADX</th>
      <th data-tip="Close > 50-day MA">LT↑</th>
      <th data-tip="BB% rising 3 days">3d↑</th>
      <th data-tip="BB% ≥ 0.8 — strongest BB signal">≥0.8</th>
      <th data-tip="MACD line rising 3 days">3d↑</th>
      <th data-tip="MACD crossed Signal 1–5d ago, MACD ≥ 0 → STRONGEST">X+0</th>
      <th data-tip="Crossed 1–5d ago, MACD < 0 but close to zero">X~0</th>
      <th data-tip="Crossed 1–5d ago, MACD far below 0 — muted">X↓</th>
      <th data-tip="Crossover 6–15d ago — sustained / decaying">X 6-15d</th>
      <th data-tip="ADX ≥ 18 — trend strong enough">≥18</th>
      <th data-tip="DI+ turning up">DI+↑</th>
      <th data-tip="DI gap widening ≥2 consecutive days (within last 5d)">Gap↑</th>
      <th data-tip="ATR-based stop loss (close − 2×ATR14)">Stop</th>
      <th></th>
    </tr>`
};
const COLS_SB = {
  widths: [180,88,64,110,70,70,68,72,108,82],
  thead: `
    <tr class="data-hd">
      <th>Ticker</th><th>Close</th><th data-tip="Day-over-day Δ in score">Δ</th>
      <th>Category</th><th>Score</th>
      <th>BB%</th><th>ADX</th>
      <th data-tip="Days since MACD crossover — lower = fresher">X age</th>
      <th data-tip="ATR-based stop loss">Stop</th>
      <th></th>
    </tr>`
};

function renderHeader() {
  const spec = TAB === "wl" ? COLS_WL : COLS_SB;
  const cols = spec.widths.map(w => `<col style="width:${w}px">`).join("");
  document.getElementById("cgHdr").innerHTML  = cols;
  document.getElementById("cgBody").innerHTML = cols;
  document.getElementById("thWrap").innerHTML = spec.thead;
  // Toggle scan-specific CSS class
  document.getElementById("hdrTbl").classList.toggle("scan",  TAB === "sb");
  document.getElementById("bodyTbl").classList.toggle("scan", TAB === "sb");
}

function currentRows() {
  if (TAB === "wl") {
    return watchlist.map(t => BY_T[t]).filter(Boolean);
  }
  // scan tab: all non-watchlist stocks with score ≥ sbMin
  const inWL = new Set(watchlist);
  const minSc = Number(document.getElementById("sbMin")?.value || 70);
  return DATA
    .filter(d => !inWL.has(d.t))
    .filter(d => V(d).sc >= minSc);
}

function render() {
  const q    = document.getElementById("q").value.trim().toLowerCase();
  const c    = document.getElementById("cat").value;
  const sec  = document.getElementById("sector").value;
  const mode = document.getElementById("sort").value;

  let rows = currentRows();
  rows = rows.filter(d => {
    const v = V(d);
    if (q && !d.t.toLowerCase().includes(q) && !d.n.toLowerCase().includes(q)) return false;
    if (c && v.cat !== c) return false;
    if (sec && (d.ind || "Unknown") !== sec) return false;
    return true;
  });

  rows.sort((a,b) => {
    if (mode === "ticker")     return a.t.localeCompare(b.t);
    if (mode === "delta-desc") return (b.dlt ?? -999) - (a.dlt ?? -999);
    const va = V(a).sc, vb = V(b).sc;
    return mode === "score-asc" ? va - vb : vb - va;
  });

  const rowFn = TAB === "wl" ? rowHtmlWL : rowHtmlSB;
  const colspan = TAB === "wl" ? 20 : 10;
  document.getElementById("tb").innerHTML = rows.map(rowFn).join("") ||
    `<tr><td colspan="${colspan}" style="text-align:center;padding:34px;color:#8b95a7">
      ${TAB === "wl"
        ? "No matching stocks in watchlist. Try adjusting filters or adding from the dropdown."
        : "No stocks currently meet the threshold. Try lowering Min Score."}
    </td></tr>`;

  // Counts + tab labels
  const wlData = watchlist.map(t => BY_T[t]).filter(Boolean);
  const inWL = new Set(watchlist);
  const minSc = Number(document.getElementById("sbMin")?.value || 70);
  const sbData = DATA.filter(d => !inWL.has(d.t) && V(d).sc >= minSc);

  document.getElementById("tabWLCount").textContent = wlData.length;
  document.getElementById("tabSBCount").textContent = sbData.length;
  document.getElementById("countTag").textContent =
    `Showing ${rows.length} ${TAB === "wl" ? "of " + wlData.length + " in watchlist" : "scan hits"} · ${DATA.length} stocks scanned · ${TF === "w" ? "weekly" : "daily"}`;
}

function buildDropdown() {
  const dd = document.getElementById("addDD");
  const in_wl = new Set(watchlist);
  const avail = DATA.filter(d => !in_wl.has(d.t))
                    .sort((a,b) => a.t.localeCompare(b.t));
  dd.innerHTML = `<option value="">— Add Nifty 500 stock (${avail.length} available) —</option>` +
    avail.map(d => `<option value="${d.t}">${d.t} — ${d.n}</option>`).join("");
}

function buildSectorDropdown() {
  const s = document.getElementById("sector");
  s.innerHTML = `<option value="">All sectors</option>` +
    SECTORS.map(v => `<option value="${v.replace(/"/g,'&quot;')}">${v}</option>`).join("");
}

// -------- Breadth strip rendering --------
function renderBreadth() {
  const b = BREADTH || {};
  const chgCls = (b.nifty_chg > 0) ? "pos" : (b.nifty_chg < 0 ? "neg" : "neu");
  const chgStr = (b.nifty_chg !== null && b.nifty_chg !== undefined)
    ? `${b.nifty_chg > 0 ? "+" : ""}${b.nifty_chg}%` : "–";
  const advCls = b.day_up > b.day_dn ? "pos" : "neg";
  const wlData = (watchlist || []).map(t => BY_T[t]).filter(Boolean);
  const wlSB   = wlData.filter(d => V(d).cat === "Strong Buy").length;
  document.getElementById("breadthRow").innerHTML = `
    <div class="bcell">
      <div class="lbl">NIFTY 50</div>
      <div class="val">${b.nifty ?? "–"}</div>
      <div class="sub2 ${chgCls}">${chgStr}</div>
    </div>
    <div class="bcell">
      <div class="lbl">Adv / Dec (N500)</div>
      <div class="val ${advCls}">${b.day_up ?? 0} / ${b.day_dn ?? 0}</div>
      <div class="sub2">avg ${b.avg_chg > 0 ? "+" : ""}${b.avg_chg ?? 0}%</div>
    </div>
    <div class="bcell">
      <div class="lbl">% above 50-MA</div>
      <div class="val">${b.adv_pct ?? 0}%</div>
      <div class="sub2">${b.n_adv ?? 0} / ${b.n_total ?? 0}</div>
    </div>
    <div class="bcell">
      <div class="lbl">Strong Buys (N500)</div>
      <div class="val sb-c">${b.n_sb ?? 0}</div>
      <div class="sub2">${b.sb_pct ?? 0}% of universe</div>
    </div>
    <div class="bcell">
      <div class="lbl">Watch (N500)</div>
      <div class="val wa-c">${b.n_wa ?? 0}</div>
      <div class="sub2">building up</div>
    </div>
    <div class="bcell">
      <div class="lbl">My Watchlist</div>
      <div class="val">${wlData.length}</div>
      <div class="sub2 sb-c">${wlSB} Strong Buy</div>
    </div>
  `;
}

// -------- Tab / timeframe toggle --------
function setTab(t) {
  TAB = t;
  localStorage.setItem(LS_TAB, TAB);
  document.getElementById("tabWL").classList.toggle("active", t === "wl");
  document.getElementById("tabSB").classList.toggle("active", t === "sb");
  // Toggle control visibility (wl-only vs sb-only controls)
  document.getElementById("wlOnly").style.display = t === "wl" ? "" : "none";
  document.getElementById("addDD").style.display = t === "wl" ? "" : "none";
  document.querySelectorAll(".controls button.primary")[0].style.display = t === "wl" ? "" : "none";
  document.querySelectorAll(".controls button")[1].style.display = t === "wl" ? "" : "none"; // Reset WL
  document.getElementById("sbOnly").style.display = t === "sb" ? "" : "none";
  document.getElementById("sbScoreWrap").style.display = t === "sb" ? "inline-flex" : "none";
  renderHeader();
  render();
}
function setTf(m) {
  TF = m;
  localStorage.setItem(LS_TF, TF);
  document.getElementById("tfD").classList.toggle("on", m === "d");
  document.getElementById("tfW").classList.toggle("on", m === "w");
  renderBreadth();
  render();
}
function addTickerDirect(t) {
  if (!watchlist.includes(t)) {
    watchlist.push(t);
    saveWL(watchlist);
    buildDropdown();
    renderBreadth();
    render();
  }
}

function addTicker() {
  const v = document.getElementById("addDD").value;
  if (!v) return;
  if (!watchlist.includes(v)) {
    watchlist.push(v);
    saveWL(watchlist);
    buildDropdown();
    renderBreadth();
    render();
  }
}
function removeTicker(t) {
  watchlist = watchlist.filter(x => x !== t);
  saveWL(watchlist);
  buildDropdown();
  renderBreadth();
  render();
}
function resetWatchlist() {
  if (!confirm("Reset to the default AM_Watch list?")) return;
  watchlist = [...DEFAULT_WL];
  saveWL(watchlist);
  buildDropdown();
  renderBreadth();
  render();
}

// ------------------------------------------------------------------
// Settings panel wiring
// ------------------------------------------------------------------
function refreshGroupLabels() {
  document.querySelectorAll('[data-gwlbl]').forEach(el => {
    el.textContent = GW[el.getAttribute('data-gwlbl')];
  });
  const sum = GW.BB + GW.MACD + GW.ADX;
  const el = document.getElementById('gwSum');
  if (el) {
    el.textContent = "Total: " + sum + (sum === 100 ? "" : " (should be 100)");
    el.style.color = sum === 100 ? "#8b95a7" : "#fbbf24";
  }
}
function syncInputs() {
  document.querySelectorAll('input[data-sw]').forEach(inp => {
    inp.value = WEIGHTS[inp.getAttribute('data-sw')] ?? 0;
  });
  document.querySelectorAll('input[data-gw]').forEach(inp => {
    inp.value = GW[inp.getAttribute('data-gw')] ?? 0;
  });
  refreshGroupLabels();
}
function onWeightChange(e) {
  const t = e.target;
  if (t.matches('input[data-sw]')) {
    WEIGHTS[t.getAttribute('data-sw')] = Number(t.value) || 0;
  } else if (t.matches('input[data-gw]')) {
    GW[t.getAttribute('data-gw')] = Number(t.value) || 0;
  } else return;
  saveSettings();
  refreshGroupLabels();
  rescoreAll();
  render();
  const st = document.getElementById('wStatus');
  if (st) { st.innerHTML = "<b>✓ re-scored</b>"; setTimeout(()=> st.textContent="", 1500); }
}
function resetWeights() {
  WEIGHTS = {...DEFAULT_WEIGHTS};
  GW      = {...DEFAULT_GW};
  saveSettings();
  syncInputs();
  rescoreAll();
  render();
  const st = document.getElementById('wStatus');
  if (st) { st.innerHTML = "<b>✓ reset to defaults</b>"; setTimeout(()=> st.textContent="", 2000); }
}
document.addEventListener('change', onWeightChange);
document.addEventListener('input',  onWeightChange);

// ------------------------------------------------------------------
// Cross-device watchlist sync via GitHub Gist
// ------------------------------------------------------------------
function syncConfigured() {
  return !!(localStorage.getItem(LS_SYNC_GIST) && localStorage.getItem(LS_SYNC_TOKEN));
}
function setSyncBadge(state, label) {
  const el = document.getElementById("syncState");
  if (!el) return;
  el.className = "sync-state " + (state || "");
  el.textContent = label || "";
}
async function syncFetch() {
  const gid = localStorage.getItem(LS_SYNC_GIST);
  const tok = localStorage.getItem(LS_SYNC_TOKEN);
  if (!gid || !tok) return { ok:false, reason:"not-configured" };
  try {
    const r = await fetch("https://api.github.com/gists/" + gid, {
      headers: { "Authorization": "token " + tok, "Accept": "application/vnd.github+json" }
    });
    if (!r.ok) return { ok:false, reason: "HTTP " + r.status };
    const j = await r.json();
    const file = j.files && (j.files["watchlist.json"] || Object.values(j.files)[0]);
    if (!file) return { ok:false, reason:"no file in gist" };
    let body = file.content;
    if (file.truncated && file.raw_url) {
      const r2 = await fetch(file.raw_url);
      body = await r2.text();
    }
    const data = JSON.parse(body || "[]");
    // Accept either a plain array or { watchlist:[...], ts:123 }
    if (Array.isArray(data)) return { ok:true, wl:data, ts:0 };
    return { ok:true, wl:(data.watchlist || []), ts:(data.ts || 0) };
  } catch (e) {
    return { ok:false, reason: e.message };
  }
}
async function syncPush() {
  const gid = localStorage.getItem(LS_SYNC_GIST);
  const tok = localStorage.getItem(LS_SYNC_TOKEN);
  if (!gid || !tok) return { ok:false };
  const payload = { watchlist: watchlist, ts: Date.now() };
  setSyncBadge("busy", "↻ syncing…");
  try {
    const r = await fetch("https://api.github.com/gists/" + gid, {
      method: "PATCH",
      headers: {
        "Authorization": "token " + tok,
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json"
      },
      body: JSON.stringify({
        files: { "watchlist.json": { content: JSON.stringify(payload, null, 2) } }
      })
    });
    if (!r.ok) {
      setSyncBadge("err", "✗ sync failed (" + r.status + ")");
      return { ok:false };
    }
    localStorage.setItem(LS_SYNC_TS, String(payload.ts));
    const t = new Date(payload.ts);
    const hm = t.toTimeString().slice(0,5);
    setSyncBadge("ok", "✓ synced " + hm);
    return { ok:true };
  } catch (e) {
    setSyncBadge("err", "✗ sync error");
    return { ok:false };
  }
}
let _syncPushTimer = null;
function syncPushDebounced() {
  if (!syncConfigured()) return;
  clearTimeout(_syncPushTimer);
  _syncPushTimer = setTimeout(syncPush, 600);
}
async function syncPullOnLoad() {
  if (!syncConfigured()) return;
  setSyncBadge("busy", "↻ pulling…");
  const res = await syncFetch();
  if (!res.ok) {
    setSyncBadge("err", "✗ sync offline");
    return;
  }
  // Remote wins on load
  watchlist = res.wl || [];
  localStorage.setItem(LS_KEY, JSON.stringify(watchlist));
  const hm = res.ts ? new Date(res.ts).toTimeString().slice(0,5) : "";
  setSyncBadge("ok", "✓ synced" + (hm ? " " + hm : ""));
  buildDropdown();
  renderBreadth();
  render();
}

// Modal UI ----------------------------------------------------------
function openSyncDialog() {
  document.getElementById("syncGistId").value = localStorage.getItem(LS_SYNC_GIST) || "";
  document.getElementById("syncToken").value  = localStorage.getItem(LS_SYNC_TOKEN) || "";
  const msg = document.getElementById("syncMsg");
  msg.className = "msg";
  msg.textContent = syncConfigured()
    ? "Sync is currently ENABLED. Leave fields as-is and click Cancel, or update them."
    : "Sync is currently disabled.";
  document.getElementById("syncDisableBtn").style.display = syncConfigured() ? "" : "none";
  document.getElementById("syncModal").classList.add("open");
}
function closeSyncDialog() {
  document.getElementById("syncModal").classList.remove("open");
}
async function saveSyncConfig() {
  const gid = document.getElementById("syncGistId").value.trim();
  const tok = document.getElementById("syncToken").value.trim();
  const msg = document.getElementById("syncMsg");
  if (!gid || !tok) {
    msg.className = "msg err";
    msg.textContent = "Both fields are required.";
    return;
  }
  localStorage.setItem(LS_SYNC_GIST, gid);
  localStorage.setItem(LS_SYNC_TOKEN, tok);
  msg.className = "msg";
  msg.textContent = "Testing connection…";
  const res = await syncFetch();
  if (!res.ok) {
    msg.className = "msg err";
    msg.textContent = "Failed: " + (res.reason || "unknown") + ". Check the Gist ID and token.";
    setSyncBadge("err", "✗ sync offline");
    return;
  }
  // If remote has an existing watchlist, offer to pull it; otherwise push local.
  if (Array.isArray(res.wl) && res.wl.length > 0) {
    const same = JSON.stringify([...res.wl].sort()) === JSON.stringify([...watchlist].sort());
    if (!same) {
      const pull = confirm(
        "Remote gist has " + res.wl.length + " tickers; local has " + watchlist.length + ".\n\n" +
        "OK  = pull remote and replace local\n" +
        "Cancel = push local and overwrite remote"
      );
      if (pull) {
        watchlist = res.wl;
        localStorage.setItem(LS_KEY, JSON.stringify(watchlist));
        buildDropdown();
        renderBreadth();
        render();
      } else {
        await syncPush();
      }
    }
  } else {
    // Remote empty — push our local
    await syncPush();
  }
  msg.className = "msg ok";
  msg.textContent = "✓ Sync enabled. Watchlist is now shared across devices.";
  setSyncBadge("ok", "✓ synced");
  document.getElementById("syncDisableBtn").style.display = "";
  setTimeout(closeSyncDialog, 900);
}
function disableSync() {
  if (!confirm("Disable sync on THIS device? Your remote gist is untouched; other devices keep syncing.")) return;
  localStorage.removeItem(LS_SYNC_GIST);
  localStorage.removeItem(LS_SYNC_TOKEN);
  localStorage.removeItem(LS_SYNC_TS);
  setSyncBadge("", "");
  document.getElementById("syncMsg").textContent = "Sync disabled on this device.";
  document.getElementById("syncDisableBtn").style.display = "none";
}

syncInputs();
buildDropdown();
buildSectorDropdown();
renderBreadth();
// Restore tab + timeframe state
document.getElementById("tfD").classList.toggle("on", TF === "d");
document.getElementById("tfW").classList.toggle("on", TF === "w");
setTab(TAB);   // this calls renderHeader + render
// Strong Buy minimum score listener
document.getElementById("sbMin").addEventListener("input", () => render());

// Pull remote watchlist on load if sync is configured
syncPullOnLoad();
</script>
</body></html>"""

    html = (html
        .replace("__STAMP_HUMAN__", STAMP_HUMAN)
        .replace("__STAMP__", STAMP)
        .replace("__DATA__", payload_json)
        .replace("__WATCHLIST__", watchlist_json)
        .replace("__BREADTH__", breadth_json)
        .replace("__SECTORS__", sectors_json))
    path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    universe = load_nifty500()
    print(f"Nifty 500 universe: {len(universe)} tickers")
    df, data, breadth = scan_universe(universe)

    ok_n = (df["status"] == "ok").sum()
    print(f"Scanned OK: {ok_n}/{len(df)}")
    print(f"Breadth: NIFTY {breadth.get('nifty')} "
          f"({breadth.get('nifty_chg')}%) · "
          f"adv {breadth.get('day_up')}/{breadth.get('day_dn')} · "
          f"SB {breadth.get('n_sb')} ({breadth.get('sb_pct')}%)")

    # Save today's snapshot for tomorrow's delta
    save_snapshot(df)

    xlsx_path  = OUT_DIR / f"AM_Watch_Scanner_{STAMP}.xlsx"
    html_path  = OUT_DIR / f"AM_Watch_Scanner_{STAMP}.html"
    index_path = OUT_DIR / "index.html"                # GH Pages entry point
    write_excel(df, xlsx_path, DEFAULT_WATCHLIST)
    write_html(df,  html_path, DEFAULT_WATCHLIST, breadth)
    # Mirror dated file to index.html so GH Pages serves it at root URL
    index_path.write_bytes(html_path.read_bytes())

    df_wl = df[df["ticker"].isin(DEFAULT_WATCHLIST) & (df["status"] == "ok")
               ].sort_values("score", ascending=False)
    print("\nAM_Watch — top by score:")
    for _, r in df_wl.head(10).iterrows():
        print(f"  {r['ticker']:12s}  {r['score']:5.1f}  {r['category']:12s}  "
              f"BB%={r['bb_pct']:.2f}  ADX={r['adx']:.1f}  DI+={r['di_plus']:.1f}")

    df_all = df[df["status"] == "ok"].sort_values("score", ascending=False)
    print(f"\nUniverse Strong Buy count: {(df_all['category']=='Strong Buy').sum()}")
    print(f"Top 10 across Nifty 500:")
    for _, r in df_all.head(10).iterrows():
        in_wl = "★" if r["ticker"] in DEFAULT_WATCHLIST else " "
        print(f"  {in_wl} {r['ticker']:12s}  {r['score']:5.1f}  {r['category']:12s}  "
              f"{r['name'][:32]}")

    print(f"\n✓ Excel: {xlsx_path}")
    print(f"✓ HTML:  {html_path}")
    print(f"✓ index: {index_path}  (GH Pages entry point)")

if __name__ == "__main__":
    main()
